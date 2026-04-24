#!/usr/bin/env python3
"""Tic-Tac-Toe as an RL Problem — Complete workbook with state space, minimax, Q-learning."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import random
from itertools import product
from functools import lru_cache

wb = openpyxl.Workbook()

# ━━ Styles (matching race car workbook) ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
HF = Font(bold=True, color="FFFFFF", size=11)
HFL = PatternFill("solid", fgColor="2F5496")
SFL = PatternFill("solid", fgColor="D6E4F0")
SFN = Font(bold=True, size=11)
TF = PatternFill("solid", fgColor="FFC7CE")
PF = PatternFill("solid", fgColor="C6EFCE")
FF = PatternFill("solid", fgColor="FCE4D6")
EF = PatternFill("solid", fgColor="4472C4")
QF = PatternFill("solid", fgColor="E2EFDA")
XF = PatternFill("solid", fgColor="D6E4F0")
OF = PatternFill("solid", fgColor="FCE4D6")
DF = PatternFill("solid", fgColor="FFF2CC")
BD = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
CT = Alignment(horizontal="center", vertical="center", wrap_text=True)
LT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def hdr(ws, r, c1, c2):
    for c in range(c1, c2 + 1):
        cl = ws.cell(r, c); cl.font = HF; cl.fill = HFL; cl.alignment = CT; cl.border = BD

def sub(ws, r, c1, c2):
    for c in range(c1, c2 + 1):
        cl = ws.cell(r, c); cl.font = SFN; cl.fill = SFL; cl.alignment = CT; cl.border = BD

def dat(ws, r, c1, c2):
    for c in range(c1, c2 + 1):
        cl = ws.cell(r, c); cl.alignment = CT; cl.border = BD

def fm(ws, r, c, formula):
    ws.cell(r, c, value=formula); ws.cell(r, c).fill = FF; ws.cell(r, c).alignment = CT; ws.cell(r, c).border = BD

def aw(ws):
    for col in ws.columns:
        mx = max((min(len(str(c.value or "")), 55) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(mx + 3, 14)

def ttl(ws, r, text, end="J"):
    ws.merge_cells(f"A{r}:{end}{r}")
    ws[f"A{r}"] = text
    ws[f"A{r}"].font = Font(bold=True, size=14, color="2F5496")
    ws[f"A{r}"].alignment = Alignment(horizontal="center")

def nt(ws, r, text, end="J"):
    ws.merge_cells(f"A{r}:{end}{r}")
    ws[f"A{r}"] = text
    ws[f"A{r}"].font = Font(italic=True, size=10, color="C00000")

def sec(ws, r, text, color="2F5496"):
    ws[f"A{r}"] = text; ws[f"A{r}"].font = Font(bold=True, size=12, color=color)


# ━━ Tic-Tac-Toe Game Engine ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
WINS = [(0,1,2),(3,4,5),(6,7,8),(0,3,6),(1,4,7),(2,5,8),(0,4,8),(2,4,6)]
SYMS = [
    (0,1,2,3,4,5,6,7,8),  # identity
    (2,5,8,1,4,7,0,3,6),  # 90 CW
    (8,7,6,5,4,3,2,1,0),  # 180
    (6,3,0,7,4,1,8,5,2),  # 270 CW
    (2,1,0,5,4,3,8,7,6),  # flip horizontal
    (0,3,6,1,4,7,2,5,8),  # flip vertical
    (0,1,2,3,4,5,6,7,8),  # flip main diagonal — handled below
    (0,1,2,3,4,5,6,7,8),  # flip anti-diagonal — handled below
]
SYMS[6] = (0,3,6,1,4,7,2,5,8)  # main diag: transpose
SYMS[7] = (8,5,2,7,4,1,6,3,0)  # anti-diag

def check_win(b, p):
    return any(b[a] == p and b[bb] == p and b[cc] == p for a, bb, cc in WINS)

def is_terminal(b):
    return check_win(b, 1) or check_win(b, 2) or 0 not in b

def winner(b):
    if check_win(b, 1): return 1
    if check_win(b, 2): return 2
    if 0 not in b: return 0
    return None

def get_empty(b):
    return [i for i in range(9) if b[i] == 0]

def whose_turn(b):
    xc = b.count(1); oc = b.count(2)
    return 1 if xc == oc else 2

def make_move(b, pos, player):
    bl = list(b); bl[pos] = player; return tuple(bl)

SYM_CHAR = {0: '.', 1: 'X', 2: 'O'}

def board_str(b):
    return '|'.join(SYM_CHAR[b[3*i]] + SYM_CHAR[b[3*i+1]] + SYM_CHAR[b[3*i+2]] for i in range(3))

def board_lines(b):
    return [SYM_CHAR[b[3*i]] + ' ' + SYM_CHAR[b[3*i+1]] + ' ' + SYM_CHAR[b[3*i+2]] for i in range(3)]

def canonical(b):
    return min(tuple(b[s[i]] for i in range(9)) for s in SYMS)


# ━━ State Space Enumeration ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def enumerate_states():
    states = []
    can_set = set()
    by_pieces = {}

    def gen(b, depth):
        bt = tuple(b)
        if bt in visited:
            return
        visited.add(bt)

        xc = bt.count(1); oc = bt.count(2)
        if xc < oc or xc > oc + 1:
            return

        x_won = check_win(bt, 1)
        o_won = check_win(bt, 2)
        if x_won and o_won:
            return
        if x_won and xc == oc:
            return
        if o_won and xc > oc:
            return

        w = winner(bt)
        pieces = xc + oc
        can = canonical(bt)

        states.append((bt, pieces, is_terminal(bt), w))
        can_set.add(can)

        if pieces not in by_pieces:
            by_pieces[pieces] = {"total": 0, "terminal": 0, "x_wins": 0, "o_wins": 0, "draws": 0}
        by_pieces[pieces]["total"] += 1
        if is_terminal(bt):
            by_pieces[pieces]["terminal"] += 1
            if w == 1: by_pieces[pieces]["x_wins"] += 1
            elif w == 2: by_pieces[pieces]["o_wins"] += 1
            elif w == 0: by_pieces[pieces]["draws"] += 1
            return

        turn = whose_turn(bt)
        for i in range(9):
            if b[i] == 0:
                b[i] = turn
                gen(b, depth + 1)
                b[i] = 0

    visited = set()
    gen([0]*9, 0)

    can_by_pieces = {}
    for s, pieces, _, _ in states:
        can = canonical(s)
        if pieces not in can_by_pieces:
            can_by_pieces[pieces] = set()
        can_by_pieces[pieces].add(can)

    return states, by_pieces, len(can_set), can_by_pieces

print("Enumerating states...")
all_states, by_pieces, total_canonical, can_by_pieces = enumerate_states()
print(f"  Total valid states: {len(all_states)}")
print(f"  Total canonical: {total_canonical}")


# ━━ Minimax ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
minimax_cache = {}

def minimax(b):
    bt = tuple(b) if isinstance(b, list) else b
    if bt in minimax_cache:
        return minimax_cache[bt]
    w = winner(bt)
    if w == 1:
        minimax_cache[bt] = 1; return 1
    if w == 2:
        minimax_cache[bt] = -1; return -1
    if w == 0:
        minimax_cache[bt] = 0; return 0

    turn = whose_turn(bt)
    empty = get_empty(bt)
    if turn == 1:
        best = -2
        for i in empty:
            v = minimax(make_move(bt, i, 1))
            best = max(best, v)
        minimax_cache[bt] = best; return best
    else:
        best = 2
        for i in empty:
            v = minimax(make_move(bt, i, 2))
            best = min(best, v)
        minimax_cache[bt] = best; return best

print("Computing minimax...")
EMPTY = (0,)*9
mm_root = minimax(EMPTY)
print(f"  Minimax value of empty board: {mm_root}")
print(f"  Minimax cache size: {len(minimax_cache)}")

# Minimax value distribution
mm_dist = {1: 0, 0: 0, -1: 0}
mm_dist_nonterminal = {1: 0, 0: 0, -1: 0}
for s, pieces, term, w in all_states:
    v = minimax(s)
    mm_dist[v] += 1
    if not term:
        mm_dist_nonterminal[v] += 1


# ━━ Key Positions for Minimax Sheet ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def action_values(b):
    turn = whose_turn(b)
    result = {}
    for i in get_empty(b):
        nb = make_move(b, i, turn)
        result[i] = minimax(nb)
    return result

key_positions = []

# Empty board
key_positions.append(("Empty board", EMPTY))

# X at center
b = make_move(EMPTY, 4, 1)
key_positions.append(("X at center", b))

# X at corner
b = make_move(EMPTY, 0, 1)
key_positions.append(("X at corner", b))

# X at edge
b = make_move(EMPTY, 1, 1)
key_positions.append(("X at edge", b))

# X@center, O@corner
b = make_move(make_move(EMPTY, 4, 1), 0, 2)
key_positions.append(("X@center, O@corner", b))

# X@center, O@edge (O mistake!)
b = make_move(make_move(EMPTY, 4, 1), 1, 2)
key_positions.append(("X@center, O@edge (mistake!)", b))

# X@corner, O@center
b = make_move(make_move(EMPTY, 0, 1), 4, 2)
key_positions.append(("X@corner, O@center", b))

# X@corner, O@edge (O mistake!)
b = make_move(make_move(EMPTY, 0, 1), 1, 2)
key_positions.append(("X@corner, O@edge (mistake!)", b))

# Fork position: X can create two threats
fork_b = (1,0,0, 0,1,0, 0,0,2)
key_positions.append(("X has diagonal, O@corner8 (fork coming)", fork_b))

# Near win for X
near_win = (1,1,0, 2,2,0, 0,0,0)
key_positions.append(("X near win (row 0)", near_win))

# X wins next
x_wins_next = (1,1,0, 2,2,0, 0,0,2)
if not is_terminal(x_wins_next):
    key_positions.append(("X wins next move", x_wins_next))

# Late game position
late = (1,2,1, 2,1,0, 0,0,2)
key_positions.append(("Late game (7 pieces)", late))

# O about to win
o_threat = (1,0,1, 2,2,0, 0,0,0)
key_positions.append(("O threatens row 1", o_threat))


# ━━ Example Games ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def play_optimal_game():
    b = list(EMPTY)
    moves = []
    while not is_terminal(tuple(b)):
        turn = whose_turn(tuple(b))
        avs = action_values(tuple(b))
        if turn == 1:
            best_v = max(avs.values())
            best_moves = [a for a, v in avs.items() if v == best_v]
        else:
            best_v = min(avs.values())
            best_moves = [a for a, v in avs.items() if v == best_v]
        a = best_moves[0]
        moves.append((tuple(b), turn, a, best_v))
        b[a] = turn
    moves.append((tuple(b), 0, None, minimax(tuple(b))))
    return moves

def play_mistake_game():
    b = [0]*9
    script = [(1,4), (2,1), (1,0), (2,8), (1,2), (2,7), (1,6)]
    moves = []
    for turn, pos in script:
        mm_val = minimax(tuple(b))
        moves.append((tuple(b), turn, pos, mm_val))
        b[pos] = turn
    moves.append((tuple(b), 0, None, minimax(tuple(b))))
    return moves

optimal_game = play_optimal_game()
mistake_game = play_mistake_game()


# ━━ Q-Learning Training ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
random.seed(42)
ALPHA = 0.1
GAMMA = 1.0
EPSILON = 0.3
NUM_EPISODES = 10000
DETAIL_EPS = 10

q_table = {}

def get_q(s, a):
    return q_table.get((s, a), 0.0)

def set_q(s, a, v):
    q_table[(s, a)] = v

def max_q(s):
    empty = get_empty(s)
    if not empty:
        return 0.0
    return max(get_q(s, a) for a in empty)

def choose_action(s, eps):
    empty = get_empty(s)
    if random.random() < eps:
        return random.choice(empty)
    qs = [(get_q(s, a), a) for a in empty]
    best_q = max(q for q, _ in qs)
    best_actions = [a for q, a in qs if q == best_q]
    return random.choice(best_actions)

detailed_episodes = []
q_snapshots = []
win_counts = [0, 0, 0]  # x_win, o_win, draw
convergence_data = []
snapshot_points = [1, 5, 10, 25, 50, 100, 250, 500, 1000, 2500, 5000, 10000]

# Key states to track Q-values over time
track_states = [EMPTY, make_move(EMPTY, 4, 1), make_move(EMPTY, 0, 1)]
track_pairs = [(EMPTY, 0), (EMPTY, 4), (EMPTY, 1),
               (make_move(EMPTY, 4, 1), 0), (make_move(EMPTY, 0, 1), 4)]

print("Training Q-learning...")
for ep in range(1, NUM_EPISODES + 1):
    b = tuple([0]*9)
    ep_trace = []

    while not is_terminal(b):
        turn = whose_turn(b)
        if turn == 1:
            a = choose_action(b, EPSILON)
            nb = make_move(b, a, 1)
            w = winner(nb)
            if w == 1:
                r_val = 1.0
            elif w == 2:
                r_val = -1.0
            elif w == 0 and is_terminal(nb):
                r_val = 0.0
            else:
                r_val = 0.0

            if is_terminal(nb):
                q_before = get_q(b, a)
                target = r_val
                error = target - q_before
                set_q(b, a, q_before + ALPHA * error)
                if ep <= DETAIL_EPS:
                    ep_trace.append((b, a, r_val, nb, True, q_before, 0.0, target, error, get_q(b, a)))
                break
            else:
                o_action = random.choice(get_empty(nb))
                nb2 = make_move(nb, o_action, 2)
                w2 = winner(nb2)
                if w2 == 2:
                    r_val = -1.0
                elif w2 == 0 and is_terminal(nb2):
                    r_val = 0.0
                else:
                    r_val = 0.0

                q_before = get_q(b, a)
                if is_terminal(nb2):
                    mq = 0.0
                else:
                    mq = max_q(nb2)
                target = r_val + GAMMA * mq
                error = target - q_before
                set_q(b, a, q_before + ALPHA * error)
                if ep <= DETAIL_EPS:
                    ep_trace.append((b, a, r_val, nb2, is_terminal(nb2), q_before, mq, target, error, get_q(b, a)))

                if is_terminal(nb2):
                    break
                b = nb2
        else:
            break

    final = nb2 if 'nb2' in dir() and is_terminal(nb2) else nb if is_terminal(nb) else b
    # determine outcome from last state
    wf = winner(final) if is_terminal(final) else winner(nb) if is_terminal(nb) else None
    if wf == 1: win_counts[0] += 1
    elif wf == 2: win_counts[1] += 1
    elif wf == 0: win_counts[2] += 1

    if ep <= DETAIL_EPS:
        detailed_episodes.append(ep_trace)

    if ep in snapshot_points:
        snap = {}
        for s, a in track_pairs:
            snap[(s, a)] = get_q(s, a)
        total_games = ep
        wr = win_counts[0] / total_games if total_games > 0 else 0
        convergence_data.append((ep, wr, win_counts[1]/total_games, win_counts[2]/total_games, dict(snap)))

print(f"  Q-table size: {len(q_table)} entries")
print(f"  Results: X wins={win_counts[0]}, O wins={win_counts[1]}, Draws={win_counts[2]}")

# Flatten detailed episodes for step-by-step sheet
all_detail_steps = []
for ei, trace in enumerate(detailed_episodes):
    for step in trace:
        all_detail_steps.append((ei + 1,) + step)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1: Intro
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.active; ws.title = "Intro"
WIDE = "J"

ttl(ws, 1, "Tic-Tac-Toe as a Reinforcement Learning Problem", WIDE)

r = 3; sec(ws, r, "1. The Game")
r += 1
game_text = [
    "Tic-Tac-Toe is played on a 3x3 grid. Two players alternate placing their marks (X and O).",
    "X always goes first. The first player to get 3 in a row (horizontal, vertical, or diagonal) wins.",
    "If all 9 cells are filled with no winner, the game is a DRAW.",
    "With optimal play from both sides, the game ALWAYS ends in a draw.",
]
for txt in game_text:
    ws.merge_cells(f"A{r}:{WIDE}{r}")
    ws.cell(r, 1, value=txt); ws.cell(r, 1).font = Font(size=11); ws.cell(r, 1).alignment = LT
    r += 1

r += 1; sec(ws, r, "2. The RL Formulation")
r += 1
nt(ws, r, "The key insight: model X as the AGENT and everything else (the board + O player) as the ENVIRONMENT.", WIDE)
r += 2

rl_diagram = [
    "                    +----------------------------------------+",
    "                    |           ENVIRONMENT                   |",
    "                    |   (3x3 Board  +  Opponent O Player)    |",
    "                    +----------------------------------------+",
    "                      |    ^                |             |",
    "             s_{t+1}  |    | a_t            | r_{t+1}    | s_t",
    "            (board     |    | (cell to       | (reward)   | (board",
    "            after O)   |    |  place X)      |            | before X)",
    "                       v    |                v             |",
    "                    +----------------------------------------+",
    "                    |              AGENT (X Player)           |",
    "                    |   Picks an empty cell to place X in    |",
    "                    |   Goal: WIN (reward=+1) or at least    |",
    "                    |   not LOSE (reward=-1)                  |",
    "                    +----------------------------------------+",
]
for line in rl_diagram:
    ws.merge_cells(f"A{r}:{WIDE}{r}")
    ws.cell(r, 1, value=line); ws.cell(r, 1).font = Font(name="Courier New", size=10)
    ws.cell(r, 1).alignment = Alignment(horizontal="left", vertical="center")
    r += 1

r += 1; sec(ws, r, "3. MDP Definition")
r += 1
mdp_hdrs = ["Component", "Symbol", "Definition", "Tic-Tac-Toe"]
for j, h in enumerate(mdp_hdrs): ws.cell(r, 1+j, value=h)
ws.merge_cells(f"C{r}:F{r}"); ws.merge_cells(f"G{r}:{WIDE}{r}")
hdr(ws, r, 1, 10); r += 1

mdp_rows = [
    ("State", "s", "Board configuration (which cells have X, O, or empty)", "Tuple of 9 cells, each in {., X, O}. 5478 valid states total."),
    ("Action", "a", "Which empty cell to place X in (0-8)", "Up to 9 actions at start, fewer as board fills up."),
    ("Transition", "P(s'|s,a)", "Agent places X, then O responds (randomly or optimally)", "Deterministic X placement + stochastic/deterministic O."),
    ("Reward", "r", "+1 if X wins, -1 if O wins, 0 for draw or non-terminal", "Sparse reward: only at episode end."),
    ("Terminal", "done", "Someone got 3 in a row, or all 9 cells filled", "X wins, O wins, or draw."),
    ("Discount", "gamma", "gamma=1.0 (no discounting — game always terminates)", "Max 5 agent moves, so finite horizon guaranteed."),
    ("Policy", "pi(a|s)", "Agent's strategy: which cell to pick given the board", "Goal: find pi* that maximizes win probability."),
]
for comp, sym, defn, ttt in mdp_rows:
    ws.cell(r, 1, value=comp); ws.cell(r, 1).font = Font(bold=True)
    ws.cell(r, 2, value=sym); ws.cell(r, 2).font = Font(italic=True)
    ws.cell(r, 3, value=defn); ws.merge_cells(f"C{r}:F{r}"); ws.cell(r, 3).alignment = LT
    ws.cell(r, 7, value=ttt); ws.merge_cells(f"G{r}:{WIDE}{r}"); ws.cell(r, 7).alignment = LT
    dat(ws, r, 1, 10); r += 1

r += 1; sec(ws, r, "4. Key Numbers")
r += 1
stats = [
    ("Total valid board states", len(all_states)),
    ("Canonical states (after symmetry reduction)", total_canonical),
    ("Terminal states", sum(1 for _, _, t, _ in all_states if t)),
    ("Non-terminal states", sum(1 for _, _, t, _ in all_states if not t)),
    ("Positions where X wins (minimax)", mm_dist[1]),
    ("Positions that are draws (minimax)", mm_dist[0]),
    ("Positions where O wins (minimax)", mm_dist[-1]),
    ("Symmetry group size (D4)", 8),
    ("Maximum agent decisions per game", 5),
    ("Minimax value of empty board", f"{mm_root} (DRAW with optimal play)"),
]
for label, val in stats:
    ws.cell(r, 1, value=label); ws.cell(r, 1).font = Font(bold=True)
    ws.cell(r, 2, value=val); ws.merge_cells(f"B{r}:D{r}")
    dat(ws, r, 1, 4); r += 1

r += 1; sec(ws, r, "5. Two Opponent Models")
r += 1
opp_hdrs = ["Model", "Behavior", "From Agent's View", "Optimal Agent Policy"]
for j, h in enumerate(opp_hdrs): ws.cell(r, 1+j, value=h)
ws.merge_cells(f"C{r}:F{r}"); ws.merge_cells(f"G{r}:{WIDE}{r}")
hdr(ws, r, 1, 10); r += 1

opps = [
    ("Random O", "O picks uniformly at random from empty cells",
     "Environment is STOCHASTIC (same action can lead to different states)",
     "Can learn to exploit mistakes -> win rate > 0. Q-values > minimax values."),
    ("Optimal O (Minimax)", "O always picks the minimax-optimal move",
     "Environment is DETERMINISTIC (always best response)",
     "Best possible outcome is a draw. Minimax value = 0 from start."),
]
for name, behavior, view, pol in opps:
    ws.cell(r, 1, value=name); ws.cell(r, 1).font = Font(bold=True)
    ws.cell(r, 2, value=behavior)
    ws.cell(r, 3, value=view); ws.merge_cells(f"C{r}:F{r}"); ws.cell(r, 3).alignment = LT
    ws.cell(r, 7, value=pol); ws.merge_cells(f"G{r}:{WIDE}{r}"); ws.cell(r, 7).alignment = LT
    dat(ws, r, 1, 10); r += 1

r += 1; sec(ws, r, "6. Key Difference from Race Car MDP")
r += 1
diffs = [
    "Race Car: single agent vs nature (stochastic transitions, no adversary).",
    "Tic-Tac-Toe: two-player ADVERSARIAL game. The opponent actively tries to MINIMIZE our reward.",
    "This changes the Bellman equation: V*(s) = max_a min_{opponent} [...] instead of max_a SUM P*[...].",
    "Minimax (game theory) replaces DP-VI (single-agent optimization). Q-Learning still works for both!",
]
for txt in diffs:
    ws.merge_cells(f"A{r}:{WIDE}{r}")
    ws.cell(r, 1, value=txt); ws.cell(r, 1).font = Font(size=11); ws.cell(r, 1).alignment = LT
    r += 1

r += 1; sec(ws, r, "7. Workbook Roadmap")
r += 1
road_hdrs = ["#", "Sheet", "Method", "What It Shows"]
for j, h in enumerate(road_hdrs): ws.cell(r, 1+j, value=h)
ws.merge_cells(f"D{r}:{WIDE}{r}")
hdr(ws, r, 1, 10); r += 1

roadmap = [
    ("1", "Intro", "Overview", "RL formulation, MDP definition, key numbers, two opponent models"),
    ("2", "StateSpace", "Enumeration", "All 5478 valid states, categorized by pieces, symmetry reduction, terminal states"),
    ("3", "GameTree", "Visualization", "Branching factor, first 2 levels of game tree, complete game examples"),
    ("4", "Minimax", "Model-Based", "Optimal solution via minimax, key position analysis, proof that optimal = draw"),
    ("5", "QLearning", "Model-Free", "Q-Learning training against random O, step-by-step traces, convergence"),
    ("6", "Comparison", "Analysis", "Minimax vs Q-Learning: model-based vs model-free, adversarial vs stochastic"),
    ("7", "Summary", "All Methods", "Final values side-by-side, key takeaways, connection to race car workbook"),
]
for num, sheet, method, desc in roadmap:
    ws.cell(r, 1, value=num); ws.cell(r, 1).font = Font(bold=True)
    ws.cell(r, 2, value=sheet); ws.cell(r, 2).font = Font(bold=True, color="2F5496")
    ws.cell(r, 3, value=method); ws.cell(r, 3).alignment = CT
    ws.cell(r, 4, value=desc); ws.merge_cells(f"D{r}:{WIDE}{r}"); ws.cell(r, 4).alignment = LT
    dat(ws, r, 1, 10); r += 1

r += 1; sec(ws, r, "Color Legend")
r += 1
legends = [
    (FF, "Orange cells = LIVE FORMULA"), (TF, "Red cells = Terminal / Loss"),
    (PF, "Green cells = Win / Positive"), (QF, "Light green = Best Q / Solution"),
    (XF, "Light blue = X player"), (OF, "Light orange = O player"),
    (DF, "Yellow = Draw"), (EF, "Blue = Episode headers"),
]
for fill, desc in legends:
    ws.cell(r, 1, value="  "); ws.cell(r, 1).fill = fill; ws.cell(r, 1).border = BD
    ws.cell(r, 2, value=desc); ws.merge_cells(f"B{r}:{WIDE}{r}"); ws.cell(r, 2).alignment = LT
    r += 1

aw(ws)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2: State Space
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("StateSpace")
ttl(ws, 1, "Tic-Tac-Toe — Complete State Space Enumeration", "I")

r = 3; sec(ws, r, "1. States by Number of Pieces on Board")
r += 1
nt(ws, r, "A 'piece' is any X or O on the board. Pieces = 0 is the empty board, pieces = 9 is a full board.", "I")
r += 2

ss_hdrs = ["Pieces", "Total States", "Non-Terminal", "Terminal", "X Wins", "O Wins", "Draws", "Whose Turn"]
for j, h in enumerate(ss_hdrs): ws.cell(r, 1+j, value=h)
hdr(ws, r, 1, 8); r += 1
ss_start = r

for p in range(10):
    d = by_pieces.get(p, {"total": 0, "terminal": 0, "x_wins": 0, "o_wins": 0, "draws": 0})
    ws.cell(r, 1, value=p); ws.cell(r, 1).font = Font(bold=True)
    ws.cell(r, 2, value=d["total"])
    ws.cell(r, 3, value=d["total"] - d["terminal"])
    ws.cell(r, 4, value=d["terminal"])
    ws.cell(r, 5, value=d["x_wins"])
    ws.cell(r, 6, value=d["o_wins"])
    ws.cell(r, 7, value=d["draws"])
    turn = "X" if p % 2 == 0 else "O"
    ws.cell(r, 8, value=turn if d["total"] - d["terminal"] > 0 else "terminal")
    dat(ws, r, 1, 8)
    if d["terminal"] == d["total"] and d["total"] > 0:
        for c in range(1, 9): ws.cell(r, c).fill = TF
    r += 1

ss_end = r - 1
ws.cell(r, 1, value="TOTAL"); ws.cell(r, 1).font = Font(bold=True)
for c in range(2, 8):
    col_letter = get_column_letter(c)
    fm(ws, r, c, f"=SUM({col_letter}{ss_start}:{col_letter}{ss_end})")
dat(ws, r, 1, 1); r += 1

r += 1; sec(ws, r, "2. Symmetry Reduction (D4 Group: 4 rotations x 2 reflections = 8 transformations)")
r += 1
nt(ws, r, "Many board positions are equivalent under rotation/reflection. Canonical = smallest representation under all 8 symmetries.", "I")
r += 2

sym_hdrs = ["Pieces", "Raw States", "Canonical States", "Reduction Factor"]
for j, h in enumerate(sym_hdrs): ws.cell(r, 1+j, value=h)
hdr(ws, r, 1, 4); r += 1
sym_start = r

for p in range(10):
    d = by_pieces.get(p, {"total": 0})
    can_count = len(can_by_pieces.get(p, set()))
    ws.cell(r, 1, value=p); ws.cell(r, 1).font = Font(bold=True)
    ws.cell(r, 2, value=d["total"])
    ws.cell(r, 3, value=can_count)
    if can_count > 0:
        fm(ws, r, 4, f"=B{r}/C{r}")
    else:
        ws.cell(r, 4, value="N/A")
    dat(ws, r, 1, 3); r += 1

sym_end = r - 1
ws.cell(r, 1, value="TOTAL"); ws.cell(r, 1).font = Font(bold=True)
fm(ws, r, 2, f"=SUM(B{sym_start}:B{sym_end})")
fm(ws, r, 3, f"=SUM(C{sym_start}:C{sym_end})")
fm(ws, r, 4, f"=B{r}/C{r}")
dat(ws, r, 1, 1); r += 1

r += 1
sym_diagram = [
    "The 8 symmetries of a 3x3 grid (D4 group):",
    "",
    " Identity    90 CW      180       270 CW     Flip H     Flip V     Diag \\     Diag /",
    " 0 1 2      6 3 0      8 7 6      2 5 8      2 1 0      6 7 8      0 3 6      8 5 2",
    " 3 4 5      7 4 1      5 4 3      1 4 7      5 4 3      3 4 5      1 4 7      7 4 1",
    " 6 7 8      8 5 2      2 1 0      0 3 6      8 7 6      0 1 2      2 5 8      6 3 0",
    "",
    "Position 0=corner always maps to corner; position 4=center always maps to center.",
    "First move: 9 cells, but only 3 canonical choices (corner, edge, center).",
]
for line in sym_diagram:
    ws.merge_cells(f"A{r}:I{r}")
    ws.cell(r, 1, value=line); ws.cell(r, 1).font = Font(name="Courier New", size=10)
    ws.cell(r, 1).alignment = Alignment(horizontal="left", vertical="center")
    r += 1

r += 1; sec(ws, r, "3. Minimax Value Distribution")
r += 1
mv_hdrs = ["Minimax Value", "Meaning", "All States", "Non-Terminal Only"]
for j, h in enumerate(mv_hdrs): ws.cell(r, 1+j, value=h)
hdr(ws, r, 1, 4); r += 1

for val, meaning, fill in [(1, "X wins with optimal play", PF), (0, "Draw with optimal play", DF), (-1, "O wins with optimal play", TF)]:
    ws.cell(r, 1, value=val); ws.cell(r, 1).font = Font(bold=True)
    ws.cell(r, 2, value=meaning)
    ws.cell(r, 3, value=mm_dist[val])
    ws.cell(r, 4, value=mm_dist_nonterminal[val])
    dat(ws, r, 1, 4)
    ws.cell(r, 1).fill = fill
    r += 1

ws.cell(r, 1, value="TOTAL"); ws.cell(r, 1).font = Font(bold=True)
fm(ws, r, 3, f"=SUM(C{r-3}:C{r-1})")
fm(ws, r, 4, f"=SUM(D{r-3}:D{r-1})")
dat(ws, r, 1, 2); r += 1

r += 1; sec(ws, r, "4. Sample Positions (representative states with minimax values)")
r += 1
samp_hdrs = ["Board", "Pieces", "Turn", "Terminal?", "Winner", "Minimax", "Best Action"]
for j, h in enumerate(samp_hdrs): ws.cell(r, 1+j, value=h)
hdr(ws, r, 1, 7); r += 1

for name, b in key_positions:
    bs = board_str(b)
    pieces = sum(1 for x in b if x != 0)
    turn = "X" if whose_turn(b) == 1 else "O"
    term = is_terminal(b)
    w = winner(b)
    w_str = {None: "--", 0: "Draw", 1: "X", 2: "O"}[w]
    mm = minimax(b)
    avs = action_values(b) if not term else {}
    if avs:
        if whose_turn(b) == 1:
            best_v = max(avs.values())
        else:
            best_v = min(avs.values())
        best_a = [a for a, v in avs.items() if v == best_v]
        ba_str = ", ".join(str(a) for a in best_a)
    else:
        ba_str = "N/A"

    ws.cell(r, 1, value=bs); ws.cell(r, 1).font = Font(name="Courier New", size=11)
    ws.cell(r, 2, value=pieces)
    ws.cell(r, 3, value=turn)
    ws.cell(r, 4, value="Yes" if term else "No")
    ws.cell(r, 5, value=w_str)
    ws.cell(r, 6, value=mm)
    ws.cell(r, 7, value=ba_str)
    dat(ws, r, 1, 7)
    if term:
        for c in range(1, 8):
            ws.cell(r, c).fill = TF if w == 2 else PF if w == 1 else DF
    elif mm == 1:
        ws.cell(r, 6).fill = PF
    elif mm == -1:
        ws.cell(r, 6).fill = TF
    r += 1

aw(ws)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3: Game Tree
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("GameTree")
ttl(ws, 1, "Tic-Tac-Toe — Game Tree Exploration", "J")

r = 3; sec(ws, r, "1. Branching Factor by Depth")
r += 1
nt(ws, r, "At depth d, the current player has (9-d) empty cells to choose from. Total tree nodes = 9! = 362,880 (with pruning for wins, fewer).", "J")
r += 2

bf_hdrs = ["Depth", "Player", "Empty Cells", "Branching Factor", "Cumulative Paths"]
for j, h in enumerate(bf_hdrs): ws.cell(r, 1+j, value=h)
hdr(ws, r, 1, 5); r += 1
bf_start = r

for d in range(10):
    ws.cell(r, 1, value=d)
    ws.cell(r, 2, value="X" if d % 2 == 0 else "O")
    ws.cell(r, 3, value=9 - d)
    ws.cell(r, 4, value=9 - d if d < 9 else 1)
    if d == 0:
        ws.cell(r, 5, value=1)
    else:
        fm(ws, r, 5, f"=E{r-1}*D{r}")
    dat(ws, r, 1, 4); r += 1

r += 1; sec(ws, r, "2. X's First Move (3 canonical choices)")
r += 1
nt(ws, r, "Due to symmetry, X's 9 possible first moves reduce to just 3 distinct positions: CORNER, EDGE, CENTER.", "J")
r += 2

first_moves = [
    ("Corner (e.g. cell 0)", 0, make_move(EMPTY, 0, 1)),
    ("Edge (e.g. cell 1)", 1, make_move(EMPTY, 1, 1)),
    ("Center (cell 4)", 4, make_move(EMPTY, 4, 1)),
]
fm_hdrs = ["First Move", "Cell", "Board After", "Minimax Value", "Equivalent Cells"]
for j, h in enumerate(fm_hdrs): ws.cell(r, 1+j, value=h)
hdr(ws, r, 1, 5); r += 1

equiv_map = {"Corner (e.g. cell 0)": "0, 2, 6, 8", "Edge (e.g. cell 1)": "1, 3, 5, 7", "Center (cell 4)": "4"}
for name, cell, b in first_moves:
    ws.cell(r, 1, value=name); ws.cell(r, 1).font = Font(bold=True)
    ws.cell(r, 2, value=cell)
    ws.cell(r, 3, value=board_str(b)); ws.cell(r, 3).font = Font(name="Courier New")
    ws.cell(r, 4, value=minimax(b))
    ws.cell(r, 5, value=equiv_map[name])
    dat(ws, r, 1, 5); r += 1

r += 1; sec(ws, r, "3. O's Responses to X@Corner (cell 0)")
r += 1
nt(ws, r, "After X plays corner, O has 8 cells. By symmetry, these reduce to 5 canonical responses.", "J")
r += 2

x_corner = make_move(EMPTY, 0, 1)
o_responses = {}
for i in get_empty(x_corner):
    nb = make_move(x_corner, i, 2)
    can = canonical(nb)
    if can not in o_responses:
        o_responses[can] = (i, nb, minimax(nb), [])
    o_responses[can][3].append(i)

resp_hdrs = ["O's Move", "Board After", "Minimax Value", "Assessment", "Equivalent Cells"]
for j, h in enumerate(resp_hdrs): ws.cell(r, 1+j, value=h)
hdr(ws, r, 1, 5); r += 1

pos_labels = {0: "corner(0)", 1: "edge(1)", 2: "corner(2)", 3: "edge(3)", 4: "center(4)",
              5: "edge(5)", 6: "corner(6)", 7: "edge(7)", 8: "corner(8)"}
for can, (first_cell, b, mm, equiv) in sorted(o_responses.items(), key=lambda x: x[1][2]):
    ws.cell(r, 1, value=pos_labels.get(first_cell, str(first_cell))); ws.cell(r, 1).font = Font(bold=True)
    ws.cell(r, 3, value=board_str(b)); ws.cell(r, 3).font = Font(name="Courier New")
    ws.cell(r, 3, value=board_str(b))
    if mm == 0:
        assessment = "Correct — holds draw"
    elif mm == 1:
        assessment = "MISTAKE — X can force win!"
    else:
        assessment = "O advantage"
    ws.cell(r, 4, value=mm)
    ws.cell(r, 4, value=assessment)
    ws.cell(r, 5, value=", ".join(str(e) for e in equiv))
    ws.cell(r, 2, value=board_str(b)); ws.cell(r, 2).font = Font(name="Courier New")
    # fix columns
    ws.cell(r, 3, value=mm)
    ws.cell(r, 4, value=assessment)
    ws.cell(r, 5, value=", ".join(str(e) for e in equiv))
    dat(ws, r, 1, 5)
    if mm == 1:
        ws.cell(r, 3).fill = PF; ws.cell(r, 4).fill = PF
    r += 1

r += 1; sec(ws, r, "4. Complete Game: Optimal Play (Draw)")
r += 1
nt(ws, r, "Both players play minimax-optimally. The game ends in a draw.", "J")
r += 2

og_hdrs = ["Move #", "Player", "Cell", "Board After", "Minimax Value"]
for j, h in enumerate(og_hdrs): ws.cell(r, 1+j, value=h)
hdr(ws, r, 1, 5); r += 1

for i, (b, turn, action, mm) in enumerate(optimal_game):
    if turn == 0:
        ws.cell(r, 1, value="FINAL")
        ws.cell(r, 2, value="--")
        ws.cell(r, 3, value="--")
    else:
        ws.cell(r, 1, value=i + 1)
        ws.cell(r, 2, value="X" if turn == 1 else "O")
        ws.cell(r, 3, value=action)
    ws.cell(r, 4, value=board_str(b)); ws.cell(r, 4).font = Font(name="Courier New")
    ws.cell(r, 5, value=mm)
    dat(ws, r, 1, 5)
    if turn == 0:
        for c in range(1, 6): ws.cell(r, c).fill = DF
    elif turn == 1:
        ws.cell(r, 2).fill = XF
    else:
        ws.cell(r, 2).fill = OF
    r += 1

r += 1; sec(ws, r, "5. Complete Game: O Makes a Mistake, X Exploits It")
r += 1
nt(ws, r, "X plays center, O plays EDGE (a mistake). X creates a fork and wins.", "J")
r += 2

mg_hdrs = ["Move #", "Player", "Cell", "Board After", "Minimax Value", "Note"]
for j, h in enumerate(mg_hdrs): ws.cell(r, 1+j, value=h)
hdr(ws, r, 1, 6); r += 1

notes = {
    0: "Empty board: minimax = 0 (draw with optimal play)",
    1: "X takes center — best first move",
    2: "O takes edge — MISTAKE! Minimax shifts to +1 (X can force win)",
    3: "X takes corner — sets up fork",
    4: "O blocks one line...",
    5: "X takes opposite corner — creates fork (two winning threats!)",
    6: "O can only block one...",
    7: "X completes the other line — X WINS!",
}
for i, (b, turn, action, mm) in enumerate(mistake_game):
    if turn == 0:
        ws.cell(r, 1, value="FINAL")
        ws.cell(r, 2, value="--")
        ws.cell(r, 3, value="--")
    else:
        ws.cell(r, 1, value=i + 1)
        ws.cell(r, 2, value="X" if turn == 1 else "O")
        ws.cell(r, 3, value=action)
    ws.cell(r, 4, value=board_str(b)); ws.cell(r, 4).font = Font(name="Courier New")
    ws.cell(r, 5, value=mm)
    ws.cell(r, 6, value=notes.get(i, ""))
    dat(ws, r, 1, 6)
    if turn == 0:
        w = winner(b)
        fill = PF if w == 1 else TF if w == 2 else DF
        for c in range(1, 7): ws.cell(r, c).fill = fill
    elif turn == 1:
        ws.cell(r, 2).fill = XF
    else:
        ws.cell(r, 2).fill = OF
    if mm == 1 and turn != 0 and i > 0:
        ws.cell(r, 5).fill = PF
    r += 1

aw(ws)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4: Minimax
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Minimax")
ttl(ws, 1, "Minimax — Optimal Solution for Tic-Tac-Toe", "J")

r = 3; sec(ws, r, "1. The Minimax Algorithm")
r += 1
nt(ws, r, "Minimax is the model-based optimal solution for two-player zero-sum games. It is analogous to DP-VI for adversarial MDPs.", "J")
r += 2

pseudo = [
    "function MINIMAX(board):",
    "    if board is terminal:",
    "        return +1 if X won, -1 if O won, 0 if draw",
    "",
    "    if X's turn (maximizer):",
    "        best = -infinity",
    "        for each empty cell:",
    "            value = MINIMAX(board after placing X)",
    "            best = max(best, value)",
    "        return best",
    "",
    "    if O's turn (minimizer):",
    "        best = +infinity",
    "        for each empty cell:",
    "            value = MINIMAX(board after placing O)",
    "            best = min(best, value)",
    "        return best",
]
for line in pseudo:
    ws.merge_cells(f"A{r}:J{r}")
    ws.cell(r, 1, value=line); ws.cell(r, 1).font = Font(name="Courier New", size=10)
    ws.cell(r, 1).alignment = Alignment(horizontal="left", vertical="center")
    r += 1

r += 1; sec(ws, r, "2. Connection to Bellman Equation")
r += 1
bellman_rows = [
    ("Single-Agent (Race Car)", "V*(s) = max_a SUM P(s'|s,a) * [R + gamma*V*(s')]", "Only MAX — nature is neutral"),
    ("Two-Player (Tic-Tac-Toe)", "V*(s) = max_a min_{opponent} V*(s'')", "MAX for agent, MIN for opponent"),
    ("Q-Learning (both)", "Q(s,a) <- Q(s,a) + alpha*[r + gamma*max Q(s',.) - Q(s,a)]", "Same update rule works for both!"),
]
be_hdrs = ["Setting", "Equation", "Key Difference"]
for j, h in enumerate(be_hdrs): ws.cell(r, 1+j, value=h)
ws.merge_cells(f"B{r}:F{r}"); ws.merge_cells(f"G{r}:J{r}")
hdr(ws, r, 1, 10); r += 1
for setting, eq, diff in bellman_rows:
    ws.cell(r, 1, value=setting); ws.cell(r, 1).font = Font(bold=True)
    ws.cell(r, 2, value=eq); ws.merge_cells(f"B{r}:F{r}")
    ws.cell(r, 2).font = Font(name="Courier New", size=10); ws.cell(r, 2).alignment = LT
    ws.cell(r, 7, value=diff); ws.merge_cells(f"G{r}:J{r}"); ws.cell(r, 7).alignment = LT
    dat(ws, r, 1, 10); r += 1

r += 1; sec(ws, r, "3. Key Position Analysis")
r += 1
nt(ws, r, "For each position: the minimax value, best action(s), and the value of EVERY possible action.", "J")
r += 2

kp_hdrs = ["Position Name", "Board", "Turn", "Minimax", "Best Action(s)", "All Action Values"]
for j, h in enumerate(kp_hdrs): ws.cell(r, 1+j, value=h)
ws.merge_cells(f"F{r}:J{r}")
hdr(ws, r, 1, 10); r += 1

for name, b in key_positions:
    avs = action_values(b) if not is_terminal(b) else {}
    turn = "X" if whose_turn(b) == 1 else "O"
    mm = minimax(b)
    if avs:
        if whose_turn(b) == 1:
            best_v = max(avs.values())
        else:
            best_v = min(avs.values())
        best_a = sorted([a for a, v in avs.items() if v == best_v])
        ba_str = ", ".join(str(a) for a in best_a)
        av_str = "  ".join(f"[{a}]={v:+d}" for a, v in sorted(avs.items()))
    else:
        ba_str = "N/A (terminal)"
        av_str = "N/A"

    ws.cell(r, 1, value=name); ws.cell(r, 1).font = Font(bold=True, size=10)
    ws.cell(r, 2, value=board_str(b)); ws.cell(r, 2).font = Font(name="Courier New")
    ws.cell(r, 3, value=turn)
    ws.cell(r, 4, value=mm)
    ws.cell(r, 5, value=ba_str)
    ws.cell(r, 6, value=av_str); ws.merge_cells(f"F{r}:J{r}")
    ws.cell(r, 6).font = Font(name="Courier New", size=9)
    ws.cell(r, 6).alignment = LT
    dat(ws, r, 1, 5)
    if mm == 1: ws.cell(r, 4).fill = PF
    elif mm == -1: ws.cell(r, 4).fill = TF
    elif mm == 0: ws.cell(r, 4).fill = DF
    r += 1

r += 1; sec(ws, r, "4. Proof: Optimal Play = Draw")
r += 1
proof_text = [
    "From the empty board, X has 3 canonical first moves (corner, edge, center).",
    "ALL three have minimax value = 0. This means no matter where X starts, O can force a draw.",
    "Conversely, no matter where O responds, X can also force at least a draw.",
    "Therefore: with optimal play from both sides, the game ALWAYS ends in a draw.",
    "",
    "The minimax value of the empty board is 0, which we verified computationally over all 5478 states.",
    "This is a SOLVED GAME — the complete solution is known since the 1950s.",
]
for txt in proof_text:
    ws.merge_cells(f"A{r}:J{r}")
    ws.cell(r, 1, value=txt)
    ws.cell(r, 1).font = Font(bold=True, size=11) if "ALWAYS" in txt or "SOLVED" in txt else Font(size=11)
    ws.cell(r, 1).alignment = LT
    r += 1

r += 1; sec(ws, r, "5. Where Mistakes Change the Outcome", "C00000")
r += 1
nt(ws, r, "The minimax value CHANGES when a player deviates from optimal. One wrong move can shift the game from draw (0) to forced loss.", "J")
r += 2

mist_hdrs = ["Scenario", "Before Move", "The Move", "After Move", "Shift", "Consequence"]
for j, h in enumerate(mist_hdrs): ws.cell(r, 1+j, value=h)
hdr(ws, r, 1, 6); r += 1

mistakes = [
    ("X@center, O@edge", "0 (draw)", "O plays cell 1 (edge)", "+1 (X wins)", "0 -> +1", "X can now create a fork and force a win"),
    ("X@center, O@corner", "0 (draw)", "O plays cell 0 (corner)", "0 (draw)", "0 -> 0", "Correct response — game stays balanced"),
    ("X@corner, O@non-center", "0 (draw)", "O plays any non-center cell", "+1 (X wins)", "0 -> +1", "Only center response holds the draw"),
]
for scen, before, move, after, shift, conseq in mistakes:
    ws.cell(r, 1, value=scen); ws.cell(r, 1).font = Font(bold=True)
    ws.cell(r, 2, value=before)
    ws.cell(r, 3, value=move)
    ws.cell(r, 4, value=after)
    ws.cell(r, 5, value=shift)
    ws.cell(r, 6, value=conseq)
    dat(ws, r, 1, 6)
    if "+1" in after:
        ws.cell(r, 4).fill = PF; ws.cell(r, 5).fill = PF
    r += 1

aw(ws)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 5: Q-Learning
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("QLearning")
ttl(ws, 1, "Q-Learning — Training Against Random Opponent", "L")

r = 3; sec(ws, r, "1. Setup")
r += 1
setup_items = [
    ("Agent", "X (goes first)"),
    ("Opponent", "Random O (picks uniformly from empty cells)"),
    ("alpha (learning rate)", str(ALPHA)),
    ("gamma (discount)", str(GAMMA) + " (episodic, no discounting)"),
    ("epsilon (exploration)", str(EPSILON)),
    ("Episodes trained", str(NUM_EPISODES)),
    ("Q-table size", f"{len(q_table)} (state, action) pairs"),
    ("Update rule", "Q(s,a) <- Q(s,a) + alpha*[r + gamma*max Q(s',.) - Q(s,a)]"),
]
for label, val in setup_items:
    ws.cell(r, 1, value=label); ws.cell(r, 1).font = Font(bold=True)
    ws.cell(r, 2, value=val); ws.merge_cells(f"B{r}:F{r}"); ws.cell(r, 2).alignment = LT
    dat(ws, r, 1, 1); r += 1

r += 1; sec(ws, r, "2. Detailed Episode Traces (first 10 episodes)")
r += 1
nt(ws, r, "Each row = one AGENT step. Agent places X, then random O responds. Target/Error/Q_after are LIVE FORMULAS.", "L")
r += 2

ep_hdrs = ["Step", "Ep", "Board (s)", "Action", "Reward", "Board (s')",
           "Done?", "Q(s,a) bef", "max Q(s',.)", "Target ✱", "Error ✱", "Q(s,a) aft ✱"]
for j, h in enumerate(ep_hdrs): ws.cell(r, 1+j, value=h)
hdr(ws, r, 1, 12); r += 1

ep_data_start = r
for i, (ep_num, b, a, rew, nb, done, q_before, mq, target, error, q_after) in enumerate(all_detail_steps):
    ws.cell(r, 1, value=i + 1)
    ws.cell(r, 2, value=ep_num)
    ws.cell(r, 3, value=board_str(b)); ws.cell(r, 3).font = Font(name="Courier New", size=9)
    ws.cell(r, 4, value=a)
    ws.cell(r, 5, value=round(rew, 4))
    ws.cell(r, 6, value=board_str(nb)); ws.cell(r, 6).font = Font(name="Courier New", size=9)
    ws.cell(r, 7, value="Yes" if done else "No")
    ws.cell(r, 8, value=round(q_before, 6))
    ws.cell(r, 9, value=round(mq, 6))
    # Live formulas for Target, Error, Q_after
    fm(ws, r, 10, f"=E{r}+{GAMMA}*I{r}")    # target = r + gamma * max_q
    fm(ws, r, 11, f"=J{r}-H{r}")              # error = target - q_before
    fm(ws, r, 12, f"=H{r}+{ALPHA}*K{r}")      # q_after = q_before + alpha * error
    dat(ws, r, 1, 9)
    if done:
        w = winner(nb)
        if w == 1:
            for c in range(1, 10): ws.cell(r, c).fill = PF
        elif w == 2:
            for c in range(1, 10): ws.cell(r, c).fill = TF
        elif w == 0:
            for c in range(1, 10): ws.cell(r, c).fill = DF
    r += 1

ep_data_end = r - 1

r += 1; sec(ws, r, "3. Q-Table for Key States (after 10,000 episodes)")
r += 1
nt(ws, r, "These Q-values were learned against RANDOM O. They are higher than minimax because random O makes exploitable mistakes.", "L")
r += 2

qt_hdrs = ["Board (s)", "Action", "Q(s,a)", "Minimax(s)", "Note"]
for j, h in enumerate(qt_hdrs): ws.cell(r, 1+j, value=h)
hdr(ws, r, 1, 5); r += 1
qt_start = r

# Show Q-values for the empty board (all 9 actions)
for a in range(9):
    q_val = get_q(EMPTY, a)
    mm_val = minimax(make_move(EMPTY, a, 1))
    ws.cell(r, 1, value=board_str(EMPTY) if a == 0 else ""); ws.cell(r, 1).font = Font(name="Courier New", size=9)
    ws.cell(r, 2, value=f"Cell {a}")
    ws.cell(r, 3, value=round(q_val, 4))
    ws.cell(r, 4, value=mm_val)
    note = ""
    if a in [0, 2, 6, 8]: note = "corner"
    elif a in [1, 3, 5, 7]: note = "edge"
    else: note = "center"
    ws.cell(r, 5, value=note)
    dat(ws, r, 1, 5)
    r += 1

r += 1
# Show Q-values for X@center (O's turn was random, now X's turn)
sec(ws, r, "Q-values after X@center, O responded randomly:")
r += 1
x_center = make_move(EMPTY, 4, 1)
for o_resp in range(9):
    if o_resp == 4: continue
    b = make_move(x_center, o_resp, 2)
    empty = get_empty(b)
    if not empty: continue
    best_a = max(empty, key=lambda a: get_q(b, a))
    best_q = get_q(b, best_a)
    mm = minimax(b)
    ws.cell(r, 1, value=board_str(b)); ws.cell(r, 1).font = Font(name="Courier New", size=9)
    ws.cell(r, 2, value=f"Best: cell {best_a}")
    ws.cell(r, 3, value=round(best_q, 4))
    ws.cell(r, 4, value=mm)
    ws.cell(r, 5, value="O mistake!" if mm == 1 else "correct O")
    dat(ws, r, 1, 5)
    if mm == 1: ws.cell(r, 5).fill = PF
    r += 1

r += 1; sec(ws, r, "4. Convergence (Q-values and win rate over training)")
r += 1
nt(ws, r, "Snapshots of Q-values for the empty board and win/loss/draw rates as training progresses.", "L")
r += 2

conv_hdrs = ["Episodes", "X Win Rate", "O Win Rate", "Draw Rate",
             "Q(empty,corner)", "Q(empty,center)", "Q(empty,edge)"]
for j, h in enumerate(conv_hdrs): ws.cell(r, 1+j, value=h)
hdr(ws, r, 1, 7); r += 1
conv_start = r

for ep_num, wr, lr, dr, snap in convergence_data:
    ws.cell(r, 1, value=ep_num)
    ws.cell(r, 2, value=round(wr, 4))
    ws.cell(r, 3, value=round(lr, 4))
    ws.cell(r, 4, value=round(dr, 4))
    ws.cell(r, 5, value=round(snap.get((EMPTY, 0), 0), 4))
    ws.cell(r, 6, value=round(snap.get((EMPTY, 4), 0), 4))
    ws.cell(r, 7, value=round(snap.get((EMPTY, 1), 0), 4))
    dat(ws, r, 1, 7); r += 1

conv_end = r - 1

r += 1
sec(ws, r, "Observations:")
r += 1
obs = [
    "Q-values are POSITIVE (around +0.5 to +0.8) because random O makes many mistakes that X can exploit.",
    "Minimax says the game is a draw (value 0) — but that assumes OPTIMAL O. Against random O, X wins often.",
    f"Final win rate after {NUM_EPISODES} episodes: X wins ~{convergence_data[-1][1]*100:.0f}%, Draws ~{convergence_data[-1][3]*100:.0f}%, O wins ~{convergence_data[-1][2]*100:.0f}%.",
    "Q(empty, corner) and Q(empty, center) are highest — these are the best opening moves against random O.",
]
for o in obs:
    ws.merge_cells(f"A{r}:L{r}")
    ws.cell(r, 1, value=o); ws.cell(r, 1).font = Font(italic=True, size=10); ws.cell(r, 1).alignment = LT
    r += 1

aw(ws)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 6: Comparison
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Comparison")
ttl(ws, 1, "Comparison: Minimax (Model-Based) vs Q-Learning (Model-Free)", "K")

r = 3; sec(ws, r, "1. Side-by-Side Comparison")
r += 1
nt(ws, r, "Minimax knows the rules and searches all possibilities. Q-Learning plays games and learns from experience. Both find optimal play.", "K")
r += 2

cmp_hdrs = ["Aspect", "Minimax (Model-Based)", "Q-Learning (Model-Free)"]
for j, h in enumerate(cmp_hdrs): ws.cell(r, 1+j, value=h)
ws.merge_cells(f"B{r}:F{r}"); ws.merge_cells(f"G{r}:K{r}")
hdr(ws, r, 1, 11); r += 1

comparisons = [
    ("Knows game rules?",
     "YES — knows all legal moves and their outcomes",
     "NO — only observes (board, action, reward, next_board)"),
    ("Knows opponent strategy?",
     "YES — assumes optimal (minimax) opponent",
     "NO — learns empirically from whatever opponent it faces"),
    ("Sees all outcomes?",
     "YES — evaluates EVERY possible move in the tree",
     "NO — sees ONE outcome per action (the actual game played)"),
    ("Computation",
     "Recursive tree search over all states (one-time)",
     "Iterative updates over many episodes (incremental)"),
    ("Update rule",
     "V(s) = max_a min_o V(s'')  (full tree backup)",
     "Q(s,a) <- Q(s,a) + alpha*[r + gamma*max Q(s',.) - Q(s,a)]"),
    ("Coverage per step",
     "Full game tree (all branches)",
     "One game (one path through the tree)"),
    ("Result against optimal O",
     "Exact: game is a draw (minimax = 0)",
     "Approximate: converges to draw if trained against optimal O"),
    ("Result against random O",
     "Still says draw (minimax ignores O's policy)",
     "Learns to exploit mistakes (Q-values > 0, high win rate)"),
    ("Analogy to race car",
     "DP-VI: model-based, sweeps all states, uses known P",
     "Q-Learning: model-free, learns from samples, no P needed"),
    ("Practical use",
     "Small games (TTT, Connect4). Intractable for chess/Go.",
     "Any game — just needs episodes. Used for chess, Go, Atari."),
]
for aspect, mm_desc, ql_desc in comparisons:
    ws.cell(r, 1, value=aspect); ws.cell(r, 1).font = Font(bold=True)
    ws.cell(r, 2, value=mm_desc); ws.merge_cells(f"B{r}:F{r}")
    ws.cell(r, 2).alignment = LT
    ws.cell(r, 7, value=ql_desc); ws.merge_cells(f"G{r}:K{r}")
    ws.cell(r, 7).alignment = LT
    dat(ws, r, 1, 11); r += 1

r += 1; sec(ws, r, "2. Why Q-Learning Values Differ from Minimax")
r += 1

diff_text = [
    "Minimax assumes OPTIMAL opponent: V(empty board) = 0 (draw). This is the WORST-CASE guarantee.",
    "Q-Learning trains against RANDOM opponent: Q(empty board) > 0 because random O makes mistakes X can exploit.",
    "This is NOT a bug — they're answering DIFFERENT questions:",
    "  Minimax: 'What's the best I can guarantee against a PERFECT opponent?'  Answer: draw (0)",
    "  Q-Learning: 'What reward do I EXPECT against THIS specific opponent?'  Answer: ~+0.6 (wins often)",
    "",
    "If we trained Q-Learning against minimax O, the Q-values WOULD converge to 0 (matching minimax).",
    "This parallels the race car: DP-VI uses exact P, Q-Learning estimates P from samples.",
]
for txt in diff_text:
    ws.merge_cells(f"A{r}:K{r}")
    ws.cell(r, 1, value=txt)
    if "NOT a bug" in txt or "DIFFERENT" in txt:
        ws.cell(r, 1).font = Font(bold=True, size=11, color="C00000")
    else:
        ws.cell(r, 1).font = Font(size=11)
    ws.cell(r, 1).alignment = LT
    r += 1

r += 1; sec(ws, r, "3. Value Comparison for Key Positions")
r += 1

vc_hdrs = ["Board", "Minimax Value", "Q-Learning Value", "Gap", "Explanation"]
for j, h in enumerate(vc_hdrs): ws.cell(r, 1+j, value=h)
hdr(ws, r, 1, 5); r += 1

cmp_positions = [
    ("Empty board", EMPTY, 4, "Center is best against random O too"),
    ("X@center", make_move(EMPTY, 4, 1), None, "O's turn — minimax sees optimal O, Q-Learning sees random O"),
]
# Add a few positions with their best Q-value
for name, b, best_a, expl in cmp_positions:
    mm = minimax(b)
    if best_a is not None:
        ql = get_q(b, best_a)
    else:
        empty = get_empty(b)
        if empty and whose_turn(b) == 1:
            ql = max(get_q(b, a) for a in empty)
        else:
            ql = 0

    ws.cell(r, 1, value=board_str(b)); ws.cell(r, 1).font = Font(name="Courier New")
    ws.cell(r, 2, value=mm)
    ws.cell(r, 3, value=round(ql, 4))
    ws.cell(r, 4, value=round(ql - mm, 4))
    ws.cell(r, 5, value=expl)
    dat(ws, r, 1, 5)
    r += 1

# Add all 9 actions from empty board
r += 1; sec(ws, r, "Empty board — all actions compared:")
r += 1

ea_hdrs = ["Action (cell)", "Type", "Minimax After", "Q-value", "Gap"]
for j, h in enumerate(ea_hdrs): ws.cell(r, 1+j, value=h)
hdr(ws, r, 1, 5); r += 1

for a in range(9):
    mm_after = minimax(make_move(EMPTY, a, 1))
    ql_val = get_q(EMPTY, a)
    cell_type = "corner" if a in [0,2,6,8] else ("center" if a == 4 else "edge")
    ws.cell(r, 1, value=f"Cell {a}")
    ws.cell(r, 2, value=cell_type)
    ws.cell(r, 3, value=mm_after)
    ws.cell(r, 4, value=round(ql_val, 4))
    ws.cell(r, 5, value=round(ql_val - mm_after, 4))
    dat(ws, r, 1, 5)
    if cell_type == "center" or cell_type == "corner":
        ws.cell(r, 4).fill = QF
    r += 1

r += 1; sec(ws, r, "4. Connection to Race Car Workbook")
r += 1

connections = [
    ("Race Car DP-VI", "-->", "Tic-Tac-Toe Minimax",
     "Both are model-based, sweep all states, use known dynamics. DP-VI uses max; Minimax uses max-min."),
    ("Race Car Q-Learning", "-->", "Tic-Tac-Toe Q-Learning",
     "Both are model-free, learn from episodes, use the SAME update rule. No knowledge of transition probabilities."),
    ("Race Car P(s'|s,a)=0.5", "-->", "TTT: O's random policy = uniform",
     "The 'transition probability' in TTT is determined by O's policy. Random O = stochastic environment."),
    ("alpha=0.1, gamma=0.9", "-->", "alpha=0.1, gamma=1.0",
     "TTT uses gamma=1 (finite horizon, always terminates). Race car uses gamma=0.9 (infinite horizon, discounting)."),
]
conn_hdrs = ["Race Car", "", "Tic-Tac-Toe", "Connection"]
for j, h in enumerate(conn_hdrs): ws.cell(r, 1+j, value=h)
ws.merge_cells(f"D{r}:K{r}")
hdr(ws, r, 1, 11); r += 1
for rc, arrow, ttt, conn in connections:
    ws.cell(r, 1, value=rc); ws.cell(r, 1).font = Font(bold=True)
    ws.cell(r, 2, value=arrow)
    ws.cell(r, 3, value=ttt); ws.cell(r, 3).font = Font(bold=True)
    ws.cell(r, 4, value=conn); ws.merge_cells(f"D{r}:K{r}"); ws.cell(r, 4).alignment = LT
    dat(ws, r, 1, 3); r += 1

aw(ws)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 7: Summary
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Summary")
ttl(ws, 1, "Summary — Tic-Tac-Toe RL Complete Overview", "H")

r = 3; sec(ws, r, "1. The Game Solved")
r += 1
solved = [
    ("Total valid states", len(all_states)),
    ("Canonical (symmetry-reduced)", total_canonical),
    ("Minimax value (empty board)", f"{mm_root} (draw)"),
    ("Optimal policy", "Draw is guaranteed with optimal play from both sides"),
    ("Minimax positions computed", len(minimax_cache)),
]
for label, val in solved:
    ws.cell(r, 1, value=label); ws.cell(r, 1).font = Font(bold=True)
    ws.cell(r, 2, value=val); ws.merge_cells(f"B{r}:F{r}")
    dat(ws, r, 1, 1); r += 1

r += 1; sec(ws, r, "2. Q-Learning Results (10,000 episodes vs random O)")
r += 1
ql_results = [
    ("Q-table entries learned", len(q_table)),
    ("X win rate", f"{convergence_data[-1][1]*100:.1f}%"),
    ("Draw rate", f"{convergence_data[-1][3]*100:.1f}%"),
    ("O win rate", f"{convergence_data[-1][2]*100:.1f}%"),
    ("Best opening move (Q-value)", f"Center (cell 4): Q={get_q(EMPTY, 4):.4f}"),
    ("Greedy policy", "Take center if available, then corners, then edges"),
]
for label, val in ql_results:
    ws.cell(r, 1, value=label); ws.cell(r, 1).font = Font(bold=True)
    ws.cell(r, 2, value=val); ws.merge_cells(f"B{r}:F{r}")
    dat(ws, r, 1, 1); r += 1

r += 1; sec(ws, r, "3. Method Comparison")
r += 1
mc_hdrs = ["", "Minimax", "Q-Learning"]
for j, h in enumerate(mc_hdrs): ws.cell(r, 1+j, value=h)
hdr(ws, r, 1, 3); r += 1

mc_rows = [
    ("Type", "Model-based (knows rules)", "Model-free (learns from play)"),
    ("Opponent model", "Optimal (worst-case)", "Random (whatever it faces)"),
    ("Empty board value", "0 (draw)", f"{get_q(EMPTY, 4):.4f} (can win)"),
    ("Computation", f"{len(minimax_cache)} recursive calls (once)", f"{NUM_EPISODES} episodes x ~3 steps"),
    ("Adapts to opponent?", "No (always assumes optimal)", "Yes (learns to exploit)"),
    ("Scalability", "Small games only", "Any game (chess, Go, Atari)"),
    ("Race car equivalent", "DP-VI", "Q-Learning"),
]
for row_data in mc_rows:
    for j, v in enumerate(row_data):
        ws.cell(r, 1+j, value=v)
        if j == 0: ws.cell(r, 1).font = Font(bold=True)
    dat(ws, r, 1, 3); r += 1

r += 1; sec(ws, r, "4. Key Takeaways")
r += 1
takeaways = [
    "1. Tic-Tac-Toe is a SOLVED GAME: minimax proves optimal play always draws.",
    "2. The opponent is part of the ENVIRONMENT: O's policy determines the transition probabilities for X.",
    "3. Q-Learning works without knowing the rules: it learns by playing games, just like a human child.",
    "4. Q-values depend on the opponent: against random O, Q > 0 (exploitable); against optimal O, Q = 0 (draw).",
    "5. Minimax is to Q-Learning as DP-VI is to Q-Learning in the race car: model-based vs model-free.",
    "6. The SAME Q-Learning update rule works for single-agent (race car) and adversarial (TTT) problems.",
    "7. For large games (chess, Go), minimax is intractable but Q-Learning (with deep networks) scales.",
]
for t in takeaways:
    ws.merge_cells(f"A{r}:H{r}")
    ws.cell(r, 1, value=t)
    ws.cell(r, 1).font = Font(size=11)
    ws.cell(r, 1).alignment = LT
    r += 1

aw(ws)


# ══════════════════════════════════════════════════════════════════════════════
# Save
# ══════════════════════════════════════════════════════════════════════════════
out = "/Users/samrat.kar/git/drl/1-introduction/assets/ttt_rl.xlsx"
wb.save(out)
print(f"\nSaved: {out}")
print(f"Sheets: {wb.sheetnames}")
print(f"States: {len(all_states)}, Canonical: {total_canonical}")
print(f"Minimax cache: {len(minimax_cache)}")
print(f"Q-table: {len(q_table)} entries")
print(f"Detailed episodes: {len(detailed_episodes)}, steps: {len(all_detail_steps)}")

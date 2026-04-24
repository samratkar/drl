#!/usr/bin/env python3
"""MDP Race Car — Full DP, MC, TD solution. Every computed cell is a live Excel formula."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import random

wb = openpyxl.Workbook()

# ━━ Styles ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
HF = Font(bold=True, color="FFFFFF", size=11)
HFL = PatternFill("solid", fgColor="2F5496")
SFL = PatternFill("solid", fgColor="D6E4F0")
SFN = Font(bold=True, size=11)
TF = PatternFill("solid", fgColor="FFC7CE")
PF = PatternFill("solid", fgColor="C6EFCE")
FF = PatternFill("solid", fgColor="FCE4D6")
EF = PatternFill("solid", fgColor="4472C4")
QF = PatternFill("solid", fgColor="E2EFDA")
BD = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
CT = Alignment(horizontal="center", vertical="center", wrap_text=True)


def hdr(ws, r, c1, c2):
    for c in range(c1, c2 + 1):
        cl = ws.cell(r, c); cl.font = HF; cl.fill = HFL; cl.alignment = CT; cl.border = BD

def sub(ws, r, c1, c2):
    for c in range(c1, c2 + 1):
        cl = ws.cell(r, c); cl.font = SFN; cl.fill = SFL; cl.alignment = CT; cl.border = BD

def dat(ws, r, c1, c2):
    for c in range(c1, c2 + 1):
        cl = ws.cell(r, c); cl.alignment = CT; cl.border = BD

def fm(ws, r, c, formula):
    ws.cell(r, c, value=formula); ws.cell(r, c).fill = FF; ws.cell(r, c).alignment = CT; ws.cell(r, c).border = BD

def aw(ws):
    for col in ws.columns:
        mx = max((min(len(str(c.value or "")), 55) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(mx + 3, 14)

def ttl(ws, r, text, end="H"):
    ws.merge_cells(f"A{r}:{end}{r}")
    ws[f"A{r}"] = text
    ws[f"A{r}"].font = Font(bold=True, size=14, color="2F5496")
    ws[f"A{r}"].alignment = Alignment(horizontal="center")

def nt(ws, r, text, end="H"):
    ws.merge_cells(f"A{r}:{end}{r}")
    ws[f"A{r}"] = text
    ws[f"A{r}"].font = Font(italic=True, size=10, color="C00000")

def sec(ws, r, text, color="2F5496"):
    ws[f"A{r}"] = text; ws[f"A{r}"].font = Font(bold=True, size=12, color=color)

ITERS = 50

# ━━ Simulate Episodes ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
random.seed(42)

def sim_ep(_):
    traj, st, step = [], "Cool", 0
    while st != "Overheated" and step < 20:
        a = ("Fast" if random.random() < 0.9 else "Slow") if st == "Cool" else ("Slow" if random.random() < 0.9 else "Fast")
        if st == "Warm" and a == "Fast":
            ns, r = "Overheated", -10
        else:
            ns = random.choice(["Cool", "Warm"]); r = 2 if a == "Fast" else 1
        traj.append((step, st, a, ns, r)); st, step = ns, step + 1
    return traj

episodes = [sim_ep(i) for i in range(5)]

all_steps = []
for ei, traj in enumerate(episodes):
    for si, (t, s, a, sn, r) in enumerate(traj):
        if sn == "Overheated":
            an = "—"
        elif si + 1 < len(traj):
            an = traj[si + 1][2]
        else:
            an = "Fast" if sn == "Cool" else "Slow"
        all_steps.append((ei + 1, s, a, r, sn, an))

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 0: Intro — What Is Reinforcement Learning?
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.active; ws.title = "Intro"
LT = Alignment(horizontal="left", vertical="center", wrap_text=True)
WIDE = "J"

ttl(ws, 1, "Introduction to Reinforcement Learning", WIDE)

# ── Section 1: What Is RL? ──────────────────────────────────────────────────
r = 3; sec(ws, r, "1. What Is Reinforcement Learning?")
r += 1
intro_text = [
    "Reinforcement Learning (RL) is a computational approach to learning from INTERACTION.",
    "An AGENT takes ACTIONS in an ENVIRONMENT, receives REWARDS, and learns a POLICY to maximize cumulative reward over time.",
    "Unlike supervised learning (given correct answers) or unsupervised learning (find patterns), RL learns by TRIAL AND ERROR.",
    "The central challenge: actions have DELAYED consequences — a good action now may only pay off many steps later.",
]
for txt in intro_text:
    ws.merge_cells(f"A{r}:{WIDE}{r}")
    ws.cell(r, 1, value=txt); ws.cell(r, 1).alignment = LT
    ws.cell(r, 1).font = Font(size=11)
    r += 1

# ── Section 2: The Agent-Environment Loop (ASCII diagram) ──────────────────
r += 1; sec(ws, r, "2. The Agent-Environment Interaction Loop")
r += 1
nt(ws, r, "At each time step t, the agent observes state s_t, takes action a_t, receives reward r_{t+1}, and transitions to s_{t+1}.", WIDE)
r += 2

diagram_lines = [
    "                          +-----------------------------------+",
    "                          |           ENVIRONMENT             |",
    "                          |   (Race Car + Track Physics)      |",
    "                          +-----------------------------------+",
    "                            |    ^              |          |",
    "                   s_{t+1}  |    | a_t          | r_{t+1}  | s_t",
    "                   (next    |    | (action)     | (reward)  | (current",
    "                    state)  |    |              |           |  state)",
    "                            v    |              v           |",
    "                          +-----------------------------------+",
    "                          |             AGENT                 |",
    "                          |   Policy: pi(a|s) -> action      |",
    "                          |   Goal: maximize sum of rewards   |",
    "                          +-----------------------------------+",
    "",
    "  Time:  t=0          t=1          t=2          t=3      ...",
    "         s_0 --a_0--> s_1 --a_1--> s_2 --a_2--> s_3 --a_3--> ...",
    "                r_1          r_2          r_3          r_4",
]
for line in diagram_lines:
    ws.merge_cells(f"A{r}:{WIDE}{r}")
    ws.cell(r, 1, value=line)
    ws.cell(r, 1).font = Font(name="Courier New", size=10)
    ws.cell(r, 1).alignment = Alignment(horizontal="left", vertical="center")
    r += 1

# ── Section 3: Key Terminology ─────────────────────────────────────────────
r += 1; sec(ws, r, "3. Key Terminology")
r += 1
terms_hdrs = ["Term", "Symbol", "Definition", "In Our Race Car Problem"]
for j, h in enumerate(terms_hdrs): ws.cell(r, 1+j, value=h)
ws.merge_cells(f"C{r}:G{r}"); ws.merge_cells(f"H{r}:{WIDE}{r}")
hdr(ws, r, 1, 10); r += 1

terms = [
    ("State", "s", "A complete description of the current situation", "{Cool, Warm, Overheated}"),
    ("Action", "a", "A choice the agent can make in a given state", "{Slow, Fast}"),
    ("Reward", "r", "Immediate scalar feedback after a transition", "+1 (slow), +2 (fast), -10 (overheat)"),
    ("Transition", "P(s'|s,a)", "Probability of reaching state s' from s via action a", "0.5/0.5 non-terminal; 1.0 terminal"),
    ("Policy", "pi(a|s)", "A rule mapping states to actions (the agent's strategy)", "e.g. Fast@Cool, Slow@Warm"),
    ("Episode", "s_0,...,s_T", "A sequence from start state to terminal state", "Cool -> ... -> Overheated"),
    ("Return", "G_t", "Cumulative discounted reward from time t onward", "G_t = r_{t+1} + gamma*G_{t+1}"),
    ("Discount", "gamma", "How much future rewards are worth (0 to 1); gamma=0.9 here", "gamma=0.9: future $1 is worth $0.90 today"),
    ("State Value", "V(s)", "Expected return starting from state s, following policy pi", "V(Cool) = expected total reward from Cool"),
    ("Action Value", "Q(s,a)", "Expected return starting from s, taking a, then following pi", "Q(Cool,Fast) = exp. reward if Fast then follow pi"),
    ("Optimal Value", "V*(s), Q*(s,a)", "Best possible expected return over ALL policies", "V*(Cool)=15.5, V*(Warm)=14.5"),
    ("Optimal Policy", "pi*(a|s)", "The policy that achieves V*(s) for every state", "pi*(Cool)=Fast, pi*(Warm)=Slow"),
]
for term, sym, defn, example in terms:
    ws.cell(r, 1, value=term); ws.cell(r, 1).font = Font(bold=True)
    ws.cell(r, 2, value=sym); ws.cell(r, 2).font = Font(italic=True)
    ws.cell(r, 3, value=defn); ws.merge_cells(f"C{r}:G{r}")
    ws.cell(r, 8, value=example); ws.merge_cells(f"H{r}:{WIDE}{r}")
    dat(ws, r, 1, 10); r += 1

# ── Section 4: The Bellman Equations ────────────────────────────────────────
r += 1; sec(ws, r, "4. The Mathematical Framework: Bellman Equations")
r += 1
nt(ws, r, "These equations express a recursive relationship: the value of a state depends on immediate reward PLUS the discounted value of the next state.", WIDE)
r += 2

eq_hdrs = ["Equation", "Name", "Formula", "What It Says"]
for j, h in enumerate(eq_hdrs): ws.cell(r, 1+j, value=h)
ws.merge_cells(f"C{r}:F{r}"); ws.merge_cells(f"G{r}:{WIDE}{r}")
hdr(ws, r, 1, 10); r += 1

equations = [
    ("1", "Bellman Expectation\n(for V^pi)",
     "V^pi(s) = SUM_a pi(a|s) * SUM_{s'} P(s'|s,a) * [R(s,a,s') + gamma * V^pi(s')]",
     "Value of s under policy pi = expected [reward + discounted future value]"),
    ("2", "Bellman Expectation\n(for Q^pi)",
     "Q^pi(s,a) = SUM_{s'} P(s'|s,a) * [R(s,a,s') + gamma * SUM_{a'} pi(a'|s') * Q^pi(s',a')]",
     "Value of taking action a in state s, then following pi"),
    ("3", "Bellman Optimality\n(for V*)",
     "V*(s) = max_a SUM_{s'} P(s'|s,a) * [R(s,a,s') + gamma * V*(s')]",
     "Optimal value = best action's expected [reward + discounted optimal future]"),
    ("4", "Bellman Optimality\n(for Q*)",
     "Q*(s,a) = SUM_{s'} P(s'|s,a) * [R(s,a,s') + gamma * max_{a'} Q*(s',a')]",
     "Optimal action-value: take a, then act optimally forever after"),
    ("5", "V-Q Relationship",
     "V*(s) = max_a Q*(s,a)     and     pi*(s) = argmax_a Q*(s,a)",
     "Optimal value = best Q; optimal policy = pick the action with highest Q"),
    ("6", "Return (recursive)",
     "G_t = r_{t+1} + gamma * G_{t+1}     with     G_T = r_T (terminal)",
     "Total discounted reward from time t; computed backward from the end"),
]
for num, name, formula, desc in equations:
    ws.cell(r, 1, value=num); ws.cell(r, 1).font = Font(bold=True)
    ws.cell(r, 2, value=name); ws.cell(r, 2).font = Font(bold=True, size=10)
    ws.cell(r, 2).alignment = CT
    ws.cell(r, 3, value=formula); ws.merge_cells(f"C{r}:F{r}")
    ws.cell(r, 3).font = Font(name="Courier New", size=10)
    ws.cell(r, 3).alignment = LT
    ws.cell(r, 7, value=desc); ws.merge_cells(f"G{r}:{WIDE}{r}")
    ws.cell(r, 7).alignment = LT
    dat(ws, r, 1, 10); r += 1

# ── Section 5: Three Families of Solution Methods ──────────────────────────
r += 1; sec(ws, r, "5. Three Families of Solution Methods")
r += 1
nt(ws, r, "All three converge to V*/Q*/pi* but differ in what they require and how they learn.", WIDE)
r += 2

meth_hdrs = ["Method", "Key Idea", "Requires Model?", "Update Rule", "Learns From"]
for j, h in enumerate(meth_hdrs): ws.cell(r, 1+j, value=h)
ws.merge_cells(f"D{r}:G{r}"); ws.merge_cells(f"H{r}:{WIDE}{r}")
hdr(ws, r, 1, 10); r += 1

methods = [
    ("Dynamic\nProgramming (DP)",
     "Sweep over ALL states, apply Bellman update using known P and R",
     "YES (full model P, R)",
     "V(s) <- SUM_a pi(a|s) * SUM_{s'} P(s'|s,a) * [R + gamma*V(s')]",
     "The model itself (no interaction needed)"),
    ("Monte Carlo\n(MC)",
     "Play complete episodes, compute actual returns, average them",
     "NO (model-free)",
     "V(s) <- average of all returns G_t observed from first visits to s",
     "Complete episodes (must reach terminal state)"),
    ("Temporal\nDifference (TD)",
     "Update after EVERY step using bootstrapped estimate of future value",
     "NO (model-free)",
     "V(s) <- V(s) + alpha*[r + gamma*V(s') - V(s)]",
     "Single transitions (online, incremental)"),
]
for name, idea, model, update, learns in methods:
    ws.cell(r, 1, value=name); ws.cell(r, 1).font = Font(bold=True, size=10)
    ws.cell(r, 1).alignment = CT
    ws.cell(r, 2, value=idea); ws.cell(r, 2).alignment = LT
    ws.cell(r, 3, value=model); ws.cell(r, 3).alignment = CT
    ws.cell(r, 4, value=update); ws.merge_cells(f"D{r}:G{r}")
    ws.cell(r, 4).font = Font(name="Courier New", size=9)
    ws.cell(r, 4).alignment = LT
    ws.cell(r, 8, value=learns); ws.merge_cells(f"H{r}:{WIDE}{r}")
    ws.cell(r, 8).alignment = LT
    dat(ws, r, 1, 10); r += 1

# ── Section 6: The Race Car Problem ────────────────────────────────────────
r += 1; sec(ws, r, "6. Our Problem: The Race Car MDP")
r += 1

problem_text = [
    "A race car drives around a track. Its engine can be in one of THREE states: Cool, Warm, or Overheated (terminal = game over).",
    "At each step, the driver chooses to go Slow or Fast. Going fast earns more reward (+2 vs +1) but risks overheating.",
    "If the car is Warm and the driver chooses Fast, it ALWAYS overheats (reward = -10). Otherwise transitions are 50/50.",
    "",
    "THE QUESTION: What is the optimal driving policy that maximizes long-term cumulative discounted reward?",
]
for txt in problem_text:
    ws.merge_cells(f"A{r}:{WIDE}{r}")
    ws.cell(r, 1, value=txt)
    ws.cell(r, 1).alignment = LT
    if "THE QUESTION" in txt:
        ws.cell(r, 1).font = Font(bold=True, size=12, color="C00000")
    else:
        ws.cell(r, 1).font = Font(size=11)
    r += 1

r += 1
# State transition diagram in ASCII
sec(ws, r, "State Transition Diagram:"); r += 1
state_diag = [
    "                     a=Slow, r=+1          a=Slow, r=+1",
    "                   +-------------+        +-------------+",
    "                   |             |        |             |",
    "                   v    0.5      |  0.5   v    0.5      |",
    "              +----------+  --------->  +----------+    |",
    "              |          | <---------   |          |----+",
    "              |   COOL   |    0.5       |   WARM   |",
    "              |          | --------->   |          |",
    "              +----------+  a=Fast,r=+2 +----------+",
    "                   |    ^                     |",
    "                   |    |  0.5                 |  a=Fast (P=1.0)",
    "                   +----+                     |  r=-10",
    "                  a=Fast, r=+2                v",
    "                                       +----------+",
    "                                       |OVERHEATED|",
    "                                       | TERMINAL |",
    "                                       |  V* = 0  |",
    "                                       +----------+",
]
for line in state_diag:
    ws.merge_cells(f"A{r}:{WIDE}{r}")
    ws.cell(r, 1, value=line)
    ws.cell(r, 1).font = Font(name="Courier New", size=10)
    ws.cell(r, 1).alignment = Alignment(horizontal="left", vertical="center")
    r += 1

# ── Section 7: Answer Preview ──────────────────────────────────────────────
r += 1; sec(ws, r, "7. The Answer (solved analytically in the Analytical sheet)")
r += 1
ans_hdrs = ["", "V* (Optimal Value)", "pi* (Optimal Policy)", "Intuition"]
for j, h in enumerate(ans_hdrs): ws.cell(r, 1+j, value=h)
ws.merge_cells(f"D{r}:{WIDE}{r}")
hdr(ws, r, 1, 10); r += 1

answers = [
    ("Cool", "15.5", "Fast", "Go fast — even if you warm up, you can slow down next turn. The +2 reward is worth the risk."),
    ("Warm", "14.5", "Slow", "Go slow — going fast means certain overheating (-10). Play it safe and cool down."),
    ("Overheated", "0", "-- (terminal)", "Game over. No more rewards possible."),
]
for state, vstar, pol, intuit in answers:
    ws.cell(r, 1, value=state); ws.cell(r, 1).font = Font(bold=True, size=11)
    ws.cell(r, 2, value=vstar); ws.cell(r, 2).font = Font(bold=True, size=11, color="C00000")
    ws.cell(r, 3, value=pol); ws.cell(r, 3).font = Font(bold=True, size=11, color="00B050")
    ws.cell(r, 4, value=intuit); ws.merge_cells(f"D{r}:{WIDE}{r}")
    ws.cell(r, 4).alignment = LT
    dat(ws, r, 1, 10)
    if state == "Overheated":
        for c in range(1, 11): ws.cell(r, c).fill = TF
    r += 1

# ── Section 8: Workbook Roadmap ────────────────────────────────────────────
r += 1; sec(ws, r, "8. Workbook Roadmap: How Each Sheet Solves This Problem")
r += 1
nt(ws, r, "Each sheet below attacks the SAME race car problem using a different method. Together they show the complete RL solution pipeline.", WIDE)
r += 2

road_hdrs = ["#", "Sheet", "Method", "What It Does", "Key Output"]
for j, h in enumerate(road_hdrs): ws.cell(r, 1+j, value=h)
ws.merge_cells(f"D{r}:G{r}"); ws.merge_cells(f"H{r}:{WIDE}{r}")
hdr(ws, r, 1, 10); r += 1

roadmap = [
    ("1", "Problem", "Definition",
     "Defines the MDP: states, actions, transition probabilities P(s'|s,a), rewards R(s,a,s'), gamma=0.9",
     "Complete MDP specification"),
    ("2", "DP-Eval", "Dynamic Programming",
     "Policy Evaluation: computes V^pi for all 4 deterministic policies using iterative Bellman updates. Shows synchronous vs in-place side by side.",
     "V^pi(s) for each policy after 50 sweeps"),
    ("3", "DP-VI", "Dynamic Programming",
     "Value Iteration: finds V*, Q*, pi* by applying the Bellman OPTIMALITY equation iteratively. Synchronous vs in-place compared.",
     "V*(Cool)=15.5, V*(Warm)=14.5, pi*"),
    ("4", "Convergence", "Analysis",
     "Tracks how V* approaches the true values over 50 iterations. Shows |error|, |DV|, and contraction ratio ~gamma=0.9.",
     "Convergence speed: sync vs in-place"),
    ("5", "Comparisons", "Analysis",
     "Side-by-side comparison: DP-Eval vs DP-VI (expectation vs max), DP-VI vs Q-Learning (model-based vs model-free), and all 5 methods.",
     "Where P(s'|s,a) appears and why methods differ"),
    ("6", "Analytical", "Algebra",
     "Solves the Bellman equation EXACTLY by hand: sets up system of equations, shows circular dependency, substitutes, and solves for V*.",
     "Closed-form V*, Q* with step-by-step derivation"),
    ("7", "Summary", "All Methods",
     "Collects V^pi, V*, Q^pi, Q* for all policies in one place. Demonstrates Policy Iteration: evaluate -> improve -> converge.",
     "Complete V/Q tables + policy iteration demo"),
    ("8", "MC-Episodes", "Monte Carlo",
     "Simulates 5 episodes following an epsilon-greedy optimal policy. Computes return G_t backward for each step.",
     "Episode trajectories with discounted returns"),
    ("9", "MC-VQ", "Monte Carlo",
     "First-visit MC estimation: averages G_t returns across episodes to estimate V^MC(s) and Q^MC(s,a).",
     "V^MC, Q^MC, greedy policy from MC"),
    ("10", "TD0", "Temporal Difference",
     "TD(0) updates V(s) after EVERY transition using V(s) <- V(s) + alpha*[r + gamma*V(s') - V(s)]. Step-by-step trace.",
     "V^TD(s) after all episodes"),
    ("11", "SARSA", "Temporal Difference",
     "On-policy Q-learning: Q(s,a) <- Q(s,a) + alpha*[r + gamma*Q(s',a') - Q(s,a)]. Uses the ACTUAL next action a'.",
     "Q^SARSA(s,a) + greedy policy"),
    ("12", "QLearning", "Temporal Difference",
     "Off-policy Q-learning: Q(s,a) <- Q(s,a) + alpha*[r + gamma*max Q(s',.) - Q(s,a)]. Uses the BEST next action.",
     "Q*(s,a) estimate + optimal policy"),
]
for num, sheet, method, desc, output in roadmap:
    ws.cell(r, 1, value=num); ws.cell(r, 1).font = Font(bold=True)
    ws.cell(r, 2, value=sheet); ws.cell(r, 2).font = Font(bold=True, color="2F5496")
    ws.cell(r, 3, value=method); ws.cell(r, 3).alignment = CT
    ws.cell(r, 4, value=desc); ws.merge_cells(f"D{r}:G{r}")
    ws.cell(r, 4).alignment = LT
    ws.cell(r, 8, value=output); ws.merge_cells(f"H{r}:{WIDE}{r}")
    ws.cell(r, 8).alignment = LT
    dat(ws, r, 1, 10); r += 1

# ── Section 9: The Big Picture ─────────────────────────────────────────────
r += 1; sec(ws, r, "9. The Big Picture: From Problem to Solution")
r += 1
pipeline = [
    "DEFINE the MDP (Problem sheet)",
    "      |",
    "      v",
    "SOLVE for V* and pi* --- three paths:",
    "      |                   |                    |",
    "  [MODEL-BASED]      [MODEL-FREE]         [MODEL-FREE]",
    "  Dynamic Prog.      Monte Carlo            TD Learning",
    "  (DP-Eval, DP-VI)   (MC-Episodes, MC-VQ)  (TD0, SARSA, QLearning)",
    "  Needs: P, R        Needs: episodes        Needs: single transitions",
    "  Exact (converges)  Unbiased, high var.    Biased, low variance",
    "      |                   |                    |",
    "      +-------------------+--------------------+",
    "                          |",
    "                          v",
    "VERIFY: all methods converge to the SAME answer: pi*(Cool)=Fast, pi*(Warm)=Slow",
    "        (Analytical sheet proves it algebraically; Convergence sheet shows DP getting there)",
]
for line in pipeline:
    ws.merge_cells(f"A{r}:{WIDE}{r}")
    ws.cell(r, 1, value=line)
    ws.cell(r, 1).font = Font(name="Courier New", size=10)
    ws.cell(r, 1).alignment = Alignment(horizontal="left", vertical="center")
    r += 1

# ── Color legend ───────────────────────────────────────────────────────────
r += 1; sec(ws, r, "Color Legend Used Throughout This Workbook")
r += 1
legends = [
    (FF, "Orange cells = LIVE FORMULA (click cell -> formula bar shows computation)"),
    (TF, "Red cells = Terminal state / penalty"),
    (PF, "Green cells = Positive reward"),
    (QF, "Light green cells = Best Q-value or solution"),
    (EF, "Blue cells = Episode headers"),
    (SFL, "Light blue cells = Initialization / sub-headers"),
]
for fill, desc in legends:
    ws.cell(r, 1, value="  "); ws.cell(r, 1).fill = fill; ws.cell(r, 1).border = BD
    ws.cell(r, 2, value=desc); ws.merge_cells(f"B{r}:{WIDE}{r}")
    ws.cell(r, 2).alignment = LT
    ws.cell(r, 2).font = Font(size=10)
    r += 1

aw(ws)

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1: Problem Definition
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Problem")

ttl(ws, 1, "MDP Race Car — Problem Definition", "F")

sec(ws, 3, "States (S)")
for i, rd in enumerate([["State", "Symbol", "Terminal?"],
    ["Cool", "C", "No"], ["Warm", "W", "No"], ["Overheated", "OH", "Yes"]]):
    r = 4 + i
    for j, v in enumerate(rd): ws.cell(r, 1+j, value=v)
    (hdr if i == 0 else dat)(ws, r, 1, 3)
    if i > 0 and rd[2] == "Yes":
        for j in range(3): ws.cell(r, 1+j).fill = TF

sec(ws, 9, "Actions (A)")
for i, rd in enumerate([["Action", "Symbol"], ["Slow", "S"], ["Fast", "F"]]):
    r = 10 + i
    for j, v in enumerate(rd): ws.cell(r, 1+j, value=v)
    (hdr if i == 0 else dat)(ws, r, 1, 2)

sec(ws, 14, "Transition Probabilities P(s'|s,a)")
tp_hdrs = ["s", "a", "P(s'=C)", "P(s'=W)", "P(s'=OH)", "Sum ✱"]
for j, h in enumerate(tp_hdrs): ws.cell(15, 1+j, value=h)
hdr(ws, 15, 1, 6)
tp_data = [("Cool","Slow",0.5,0.5,0),("Cool","Fast",0.5,0.5,0),
           ("Warm","Slow",0.5,0.5,0),("Warm","Fast",0,0,1)]
for i, (s, a, pc, pw, po) in enumerate(tp_data):
    r = 16 + i
    ws.cell(r,1,value=s); ws.cell(r,2,value=a)
    ws.cell(r,3,value=pc); ws.cell(r,4,value=pw); ws.cell(r,5,value=po)
    fm(ws, r, 6, f"=C{r}+D{r}+E{r}")
    dat(ws, r, 1, 5)

sec(ws, 21, "Reward Function R(s,a,s')")
for j, h in enumerate(["Condition", "R"]): ws.cell(22, 1+j, value=h)
hdr(ws, 22, 1, 2)
for i, (cond, rv) in enumerate([("s' non-terminal, a=Slow", 1),("s' non-terminal, a=Fast", 2),("s'=Overheated", -10)]):
    r = 23 + i; ws.cell(r,1,value=cond); ws.cell(r,2,value=rv); dat(ws, r, 1, 2)

sec(ws, 27, "Parameters")
ws.cell(28,1,value="γ (gamma)"); ws.cell(28,2,value=0.9)
ws.cell(29,1,value="α (learning rate)"); ws.cell(29,2,value=0.1)
dat(ws, 28, 1, 2); dat(ws, 29, 1, 2)
aw(ws)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2: DP — Policy Evaluation (V^π)
# ══════════════════════════════════════════════════════════════════════════════
# ── Helper: write one policy eval block ──────────────────────────────────────
# sync_vc_f / sync_vw_f: synchronous formulas (both ref previous row p)
# inplace_vc_f / inplace_vw_f: in-place formulas
#   - Cool is updated first → V(Cool) uses previous row for both V(C) and V(W)
#   - Warm is updated second → V(Warm) uses CURRENT row's V(Cool) = B{r}, previous V(W) = H{p}
# Cols B-E = synchronous (B=V_C, C=V_W, D=V_OH, E=|ΔV|)
# Cols G-J = in-place   (G=V_C, H=V_W, I=V_OH, J=|ΔV|)
# Col K = converged diff

def write_pe_dual(ws, start, name, color, sync_vc, sync_vw, inp_vc, inp_vw, eq_sync, eq_inp):
    r = start
    sec(ws, r, name, color); r += 1
    # Synchronous equation
    nt(ws, r, "SYNCHRONOUS: " + eq_sync, "K"); r += 1
    # In-place equation
    nt(ws, r, "IN-PLACE: " + eq_inp, "K"); r += 1
    hdrs = ["k",
            "V_k(C) sync ✱", "V_k(W) sync ✱", "V_k(OH)", "|ΔV| sync ✱", "",
            "V_k(C) inplace ✱", "V_k(W) inplace ✱", "V_k(OH)", "|ΔV| inplace ✱",
            "Diff ✱"]
    for j, h in enumerate(hdrs): ws.cell(r, 1+j, value=h)
    hdr(ws, r, 1, 11)
    ws.cell(r, 6).fill = PatternFill()  # blank spacer col
    r += 1; init = r
    # k=0 init row
    ws.cell(r,1,value=0)
    ws.cell(r,2,value=0); ws.cell(r,3,value=0); ws.cell(r,4,value=0); ws.cell(r,5,value="—")
    ws.cell(r,6,value="")
    ws.cell(r,7,value=0); ws.cell(r,8,value=0); ws.cell(r,9,value=0); ws.cell(r,10,value="—")
    ws.cell(r,11,value="—")
    dat(ws, r, 1, 11)
    for k in range(1, ITERS + 1):
        r = init + k; p = r - 1
        ws.cell(r, 1, value=k)
        # ── Synchronous: both V(C) and V(W) ref previous row ──
        fm(ws, r, 2, sync_vc.format(vc=f"B{p}", vw=f"C{p}", voh=f"D{p}"))
        fm(ws, r, 3, sync_vw.format(vc=f"B{p}", vw=f"C{p}", voh=f"D{p}"))
        ws.cell(r, 4, value=0)
        fm(ws, r, 5, f"=MAX(ABS(B{r}-B{p}),ABS(C{r}-C{p}))")
        ws.cell(r, 6, value="")
        # ── In-place: Cool updated first (uses prev), Warm uses NEW Cool ──
        fm(ws, r, 7, inp_vc.format(vc=f"G{p}", vw=f"H{p}", voh=f"I{p}"))
        fm(ws, r, 8, inp_vw.format(vc=f"G{r}", vw=f"H{p}", voh=f"I{p}"))  # G{r} = just-updated Cool!
        ws.cell(r, 9, value=0)
        fm(ws, r, 10, f"=MAX(ABS(G{r}-G{p}),ABS(H{r}-H{p}))")
        # ── Diff: how far apart are sync vs inplace this iteration ──
        fm(ws, r, 11, f"=MAX(ABS(B{r}-G{r}),ABS(C{r}-H{r}))")
        dat(ws, r, 1, 1); dat(ws, r, 4, 4); dat(ws, r, 6, 6); dat(ws, r, 9, 9)
    return init + ITERS

ws = wb.create_sheet("DP-Eval")
ttl(ws, 1, "DP — Policy Evaluation: Synchronous vs In-Place", "K")
nt(ws, 2, "Synchronous: V_{k+1}(s) uses V_k for ALL states (full previous sweep).  In-Place: V(s) uses latest available values within the SAME sweep.", "K")
nt(ws, 3, "k = iteration (one full sweep). Both converge to the same V^π — in-place is often faster. The Bellman EQUATION describes the fixed point; these are two ALGORITHMS to reach it.", "K")

# Sync formulas: both {vc} and {vw} come from previous row
# In-place: Cool first (same as sync), Warm second → uses updated Cool from SAME row
# The formulas use named placeholders {vc}, {vw}, {voh} that get filled with cell refs

pe_caut = write_pe_dual(ws, 5, "Policy 1: Cautious (Always Slow)", "2F5496",
    sync_vc="=0.5*(1+0.9*{vc})+0.5*(1+0.9*{vw})",
    sync_vw="=0.5*(1+0.9*{vc})+0.5*(1+0.9*{vw})",
    inp_vc ="=0.5*(1+0.9*{vc})+0.5*(1+0.9*{vw})",
    inp_vw ="=0.5*(1+0.9*{vc})+0.5*(1+0.9*{vw})",  # {vc} = G{r} (just updated)
    eq_sync="V_{k+1}(C) = 0.5·(1+γ·V_k(C)) + 0.5·(1+γ·V_k(W)).  V_{k+1}(W) same.  Both from PREVIOUS sweep k.",
    eq_inp ="V(C) ← 0.5·(1+γ·V_old(C)) + 0.5·(1+γ·V_old(W)).  Then V(W) ← 0.5·(1+γ·V_NEW(C)) + 0.5·(1+γ·V_old(W)).  Cool updated first!")

pe_opt = write_pe_dual(ws, pe_caut + 3, "Policy 2: Optimal (Fast@Cool, Slow@Warm)", "00B050",
    sync_vc="=0.5*(2+0.9*{vc})+0.5*(2+0.9*{vw})",
    sync_vw="=0.5*(1+0.9*{vc})+0.5*(1+0.9*{vw})",
    inp_vc ="=0.5*(2+0.9*{vc})+0.5*(2+0.9*{vw})",
    inp_vw ="=0.5*(1+0.9*{vc})+0.5*(1+0.9*{vw})",
    eq_sync="V_{k+1}(C) = 0.5·(2+γ·V_k(C))+0.5·(2+γ·V_k(W)).  V_{k+1}(W) = 0.5·(1+γ·V_k(C))+0.5·(1+γ·V_k(W)).  Both from sweep k.",
    eq_inp ="V(C) ← fast formula with V_old.  Then V(W) ← slow formula using V_NEW(C) + V_old(W).  Warm sees Cool's updated value!")

pe_agg = write_pe_dual(ws, pe_opt + 3, "Policy 3: Aggressive (Always Fast)", "C00000",
    sync_vc="=0.5*(2+0.9*{vc})+0.5*(2+0.9*{vw})",
    sync_vw="=1*(-10+0.9*{voh})",
    inp_vc ="=0.5*(2+0.9*{vc})+0.5*(2+0.9*{vw})",
    inp_vw ="=1*(-10+0.9*{voh})",
    eq_sync="V_{k+1}(C) = 0.5·(2+γ·V_k(C))+0.5·(2+γ·V_k(W)).  V_{k+1}(W) = -10+γ·0 = -10.  (Warm→OH always.)",
    eq_inp ="V(C) ← fast formula with V_old.  V(W) ← -10.  (Warm doesn't depend on Cool here, so sync=inplace for W.)")

pe_worst = write_pe_dual(ws, pe_agg + 3, "Policy 4: Worst (Slow@Cool, Fast@Warm)", "7030A0",
    sync_vc="=0.5*(1+0.9*{vc})+0.5*(1+0.9*{vw})",
    sync_vw="=1*(-10+0.9*{voh})",
    inp_vc ="=0.5*(1+0.9*{vc})+0.5*(1+0.9*{vw})",
    inp_vw ="=1*(-10+0.9*{voh})",
    eq_sync="V_{k+1}(C) = 0.5·(1+γ·V_k(C))+0.5·(1+γ·V_k(W)).  V_{k+1}(W) = -10.  Slow@Cool, Fast@Warm.",
    eq_inp ="V(C) ← slow formula with V_old.  V(W) ← -10.  (Again W doesn't depend on C, so identical.)")

# ── Reconciliation: Direct assignment vs Error form ─────────────────────────
r = pe_worst + 3
sec(ws, r, "Why DP Doesn't Show a 'Target' and 'Error' (Reconciliation)", "C00000")
r += 1
nt(ws, r, "You may have seen the update written as: V(s) <- V(s) + alpha * [target - V(s)].  That IS what's happening here, with alpha=1.", "K")
r += 1
nt(ws, r, "When alpha=1: V(s) + 1*(target - V(s)) = V(s) + target - V(s) = target.  So 'V_new = target' and 'V_new = V_old + error' are THE SAME THING.", "K")
r += 1; r += 1

sec(ws, r, "The Two Forms Side by Side:")
r += 1
rec_hdrs = ["", "Error Form (explicit)", "", "Direct Form (what the sheet shows)", "Same?"]
for j, h in enumerate(rec_hdrs): ws.cell(r, 1+j, value=h)
hdr(ws, r, 1, 5); r += 1

# Use Policy 2 (Optimal) sync rows for demonstration — pe_caut+3 block starts optimal
# pe_opt is the last row of Optimal policy block
# We'll use a fixed example: k=1 of Cautious (pe_caut block, init at row after header)
# Cautious starts at row 5: sec=5, eq1=6, eq2=7, hdr=8, init=9, k=1 at row 10
caut_init = 9  # k=0 row for cautious sync
caut_k1 = caut_init + 1  # k=1 row
caut_k0 = caut_init       # k=0 row (previous)

rec_rows = [
    ("Step 1: V_old(s)",
     f"V_0(Cool) = 0",
     "",
     f"V_0(Cool) = 0",
     ""),
    ("Step 2: Compute target",
     "target = SUM P*[R + gamma*V_old(s')]",
     "",
     "(this IS the RHS of the formula)",
     ""),
    ("Step 3: Compute error",
     "error = target - V_old(s)",
     "",
     "(implicit: error = target - 0 = target)",
     ""),
    ("Step 4: V_new(s)",
     "V_old + alpha * error = 0 + 1.0*(target - 0) = target",
     "",
     "V_new = 0.5*(1+0.9*V_0(C)) + 0.5*(1+0.9*V_0(W)) = target",
     ""),
]
for lbl, err_form, spacer, direct, same in rec_rows:
    ws.cell(r, 1, value=lbl); ws.cell(r, 1).font = Font(bold=True)
    ws.cell(r, 2, value=err_form)
    ws.cell(r, 3, value=spacer)
    ws.cell(r, 4, value=direct)
    ws.cell(r, 5, value=same)
    dat(ws, r, 1, 5); r += 1

r += 1
sec(ws, r, "Live Formula Proof (using Cautious Policy, k=1, Cool state):")
r += 1
proof_hdrs = ["Quantity", "Formula", "Value"]
for j, h in enumerate(proof_hdrs): ws.cell(r, 1+j, value=h)
hdr(ws, r, 1, 3); r += 1

# V_old = B{caut_k0} (which is 0)
# V_new (direct) = B{caut_k1}
# target = same formula as B{caut_k1}
# error = target - V_old
# V_old + 1*error = V_old + target - V_old = target
ws.cell(r, 1, value="V_old = V_0(Cool)"); ws.cell(r, 1).font = Font(bold=True)
fm(ws, r, 2, f"=B{caut_k0}")
fm(ws, r, 3, f"=B{caut_k0}")
dat(ws, r, 1, 1); proof_vold = r; r += 1

ws.cell(r, 1, value="target = SUM P*[R+g*V_old]"); ws.cell(r, 1).font = Font(bold=True)
fm(ws, r, 2, f"=0.5*(1+0.9*B{caut_k0})+0.5*(1+0.9*C{caut_k0})")
fm(ws, r, 3, f"=0.5*(1+0.9*B{caut_k0})+0.5*(1+0.9*C{caut_k0})")
dat(ws, r, 1, 1); proof_target = r; r += 1

ws.cell(r, 1, value="error = target - V_old"); ws.cell(r, 1).font = Font(bold=True)
fm(ws, r, 2, f"=C{proof_target}-C{proof_vold}")
fm(ws, r, 3, f"=C{proof_target}-C{proof_vold}")
dat(ws, r, 1, 1); proof_error = r; r += 1

ws.cell(r, 1, value="ERROR FORM: V_old + 1.0*error"); ws.cell(r, 1).font = Font(bold=True, color="C00000")
fm(ws, r, 2, f"=C{proof_vold}+1.0*C{proof_error}")
fm(ws, r, 3, f"=C{proof_vold}+1.0*C{proof_error}")
dat(ws, r, 1, 1); proof_err_result = r; r += 1

ws.cell(r, 1, value="DIRECT FORM: V_1(Cool) from table"); ws.cell(r, 1).font = Font(bold=True, color="00B050")
fm(ws, r, 2, f"=B{caut_k1}")
fm(ws, r, 3, f"=B{caut_k1}")
dat(ws, r, 1, 1); proof_dir_result = r; r += 1

ws.cell(r, 1, value="Same result?"); ws.cell(r, 1).font = Font(bold=True, size=12)
fm(ws, r, 2, f'=IF(ABS(C{proof_err_result}-C{proof_dir_result})<0.0001,"YES — identical!","NO")')
ws.cell(r, 2).font = Font(bold=True, size=12, color="00B050")
dat(ws, r, 1, 2); r += 1

r += 1
sec(ws, r, "Why DP Uses alpha=1 But TD Uses alpha=0.1:")
r += 1
reasons = [
    ("DP (alpha=1)",
     "Target is EXACT: computed from the FULL model (all transitions, all probabilities). No noise, no sampling error.",
     "We can jump directly to the target. No need to inch — the target is correct (for the current V estimates)."),
    ("TD (alpha=0.1)",
     "Target is a NOISY SAMPLE: based on ONE random transition (s,a,r,s'). Different episodes give different targets.",
     "Must inch slowly (alpha=0.1) to average out the noise. Jumping to one sample's target would be volatile."),
    ("MC (alpha=1/N)",
     "Target is the ACTUAL return G_t from one episode. Unbiased but high variance across episodes.",
     "Averages all observed returns. Equivalent to alpha=1/N where N is the visit count."),
]
for method, target_desc, why_alpha in reasons:
    ws.cell(r, 1, value=method); ws.cell(r, 1).font = Font(bold=True, size=11)
    ws.cell(r, 2, value=target_desc)
    ws.cell(r, 3, value=why_alpha)
    dat(ws, r, 1, 3); r += 1

r += 1
nt(ws, r, "BOTTOM LINE: V_new = target (DP) and V_new = V_old + alpha*(target - V_old) (TD) are the same algorithm with different alpha.", "K")
r += 1
nt(ws, r, "DP iterates because V_k(s') in the target is itself an estimate — it takes MULTIPLE SWEEPS for these estimates to propagate and converge.", "K")

aw(ws)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3: DP — Value Iteration (V*, Q*, π*) — Synchronous vs In-Place
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("DP-VI")
ttl(ws, 1, "DP — Value Iteration: Synchronous vs In-Place", "S")
nt(ws, 2, "Synchronous: V_{k+1}(s) = max_a Σ P·[R+γ·V_k(s')] using ALL values from previous sweep k.", "S")
nt(ws, 3, "In-Place: V(s) ← max_a Σ P·[R+γ·V(s')] using latest available V — Cool updated first, Warm uses NEW Cool within same sweep.", "S")
nt(ws, 4, "Both converge to the same V*, Q*, π* — in-place often converges faster (fewer iterations for same |ΔV|).", "S")

# ── Synchronous VI ──
r = 6; sec(ws, r, "Synchronous Value Iteration", "2F5496")
r += 1
vi_hdrs = ["k","Q(C,S) ✱","Q(C,F) ✱","V*(C) ✱","Q(W,S) ✱","Q(W,F) ✱","V*(W) ✱","V*(OH)","π*(C) ✱","π*(W) ✱","|ΔV| ✱"]
for j, h in enumerate(vi_hdrs): ws.cell(r, 1+j, value=h)
hdr(ws, r, 1, 11)

r += 1; vi_init = r
ws.cell(r,1,value=0)
for c in range(2,9): ws.cell(r, c, value=0)
ws.cell(r,9,value="—"); ws.cell(r,10,value="—"); ws.cell(r,11,value="—")
dat(ws, r, 1, 11)

for k in range(1, ITERS + 1):
    r = vi_init + k; p = r - 1
    ws.cell(r, 1, value=k)
    fm(ws, r, 2, f"=0.5*(1+0.9*D{p})+0.5*(1+0.9*G{p})")       # Q(C,S)
    fm(ws, r, 3, f"=0.5*(2+0.9*D{p})+0.5*(2+0.9*G{p})")       # Q(C,F)
    fm(ws, r, 4, f"=MAX(B{r},C{r})")                             # V*(C)
    fm(ws, r, 5, f"=0.5*(1+0.9*D{p})+0.5*(1+0.9*G{p})")       # Q(W,S)
    fm(ws, r, 6, f"=1*(-10+0.9*H{p})")                           # Q(W,F)
    fm(ws, r, 7, f"=MAX(E{r},F{r})")                             # V*(W)
    ws.cell(r, 8, value=0)                                        # V*(OH)
    fm(ws, r, 9, f'=IF(C{r}>=B{r},"Fast","Slow")')              # π*(C)
    fm(ws, r, 10, f'=IF(F{r}>=E{r},"Fast","Slow")')             # π*(W)
    fm(ws, r, 11, f"=MAX(ABS(D{r}-D{p}),ABS(G{r}-G{p}))")      # |ΔV|
    dat(ws, r, 1, 1); dat(ws, r, 8, 8)

vi_last = vi_init + ITERS

# ── In-Place VI ──
r = vi_last + 3; sec(ws, r, "In-Place Value Iteration (Cool updated first, then Warm uses new Cool)", "00B050")
r += 1
ip_hdrs = ["k","Q(C,S)","Q(C,F)","V*(C)","Q(W,S)","Q(W,F)","V*(W)","V*(OH)","pi*(C)","pi*(W)","|DV|"]
for j, h in enumerate(ip_hdrs): ws.cell(r, 1+j, value=h)
hdr(ws, r, 1, 11)

r += 1; ip_init = r
ws.cell(r,1,value=0)
for c in range(2,9): ws.cell(r, c, value=0)
ws.cell(r,9,value="---"); ws.cell(r,10,value="---"); ws.cell(r,11,value="---")
dat(ws, r, 1, 11)

for k in range(1, ITERS + 1):
    r = ip_init + k; p = r - 1
    ws.cell(r, 1, value=k)
    fm(ws, r, 2, f"=0.5*(1+0.9*D{p})+0.5*(1+0.9*G{p})")
    fm(ws, r, 3, f"=0.5*(2+0.9*D{p})+0.5*(2+0.9*G{p})")
    fm(ws, r, 4, f"=MAX(B{r},C{r})")
    fm(ws, r, 5, f"=0.5*(1+0.9*D{r})+0.5*(1+0.9*G{p})")       # D{r} = just-updated Cool!
    fm(ws, r, 6, f"=1*(-10+0.9*H{p})")
    fm(ws, r, 7, f"=MAX(E{r},F{r})")
    ws.cell(r, 8, value=0)
    fm(ws, r, 9, f'=IF(C{r}>=B{r},"Fast","Slow")')
    fm(ws, r, 10, f'=IF(F{r}>=E{r},"Fast","Slow")')
    fm(ws, r, 11, f"=MAX(ABS(D{r}-D{p}),ABS(G{r}-G{p}))")
    dat(ws, r, 1, 1); dat(ws, r, 8, 8)

ip_last = ip_init + ITERS

# ── Comparison ──
r = ip_last + 2; sec(ws, r, "Converged Values Comparison")
r += 1
for j, h in enumerate(["", "V*(Cool)", "V*(Warm)", "Match?"]): ws.cell(r, 1+j, value=h)
hdr(ws, r, 1, 4); r += 1
ws.cell(r,1,value="Synchronous"); fm(ws, r, 2, f"=D{vi_last}"); fm(ws, r, 3, f"=G{vi_last}"); dat(ws, r, 1, 3); r += 1
ws.cell(r,1,value="In-Place"); fm(ws, r, 2, f"=D{ip_last}"); fm(ws, r, 3, f"=G{ip_last}"); dat(ws, r, 1, 3); r += 1
ws.cell(r,1,value="Analytical"); ws.cell(r, 2, value=15.5); ws.cell(r, 3, value=14.5); dat(ws, r, 1, 3); r += 1
ws.cell(r,1,value="Sync error"); fm(ws, r, 2, f"=ABS(B{r-3}-B{r-1})"); fm(ws, r, 3, f"=ABS(C{r-3}-C{r-1})")
fm(ws, r, 4, f'=IF(MAX(B{r},C{r})<0.001,"Converged","Not yet")'); dat(ws, r, 1, 4); r += 1
ws.cell(r,1,value="InPlace error"); fm(ws, r, 2, f"=ABS(B{r-3}-B{r-2})"); fm(ws, r, 3, f"=ABS(C{r-3}-C{r-2})")
fm(ws, r, 4, f'=IF(MAX(B{r},C{r})<0.001,"Converged","Not yet")'); dat(ws, r, 1, 4)

aw(ws)


# Convergence Tracker Sheet
ws_c = wb.create_sheet("Convergence")
ttl(ws_c, 1, "V* Convergence -- Synchronous vs In-Place approaching true V*", "N")
nt(ws_c, 2, "Analytical: V*(Cool)=15.5, V*(Warm)=14.5.  |DV| shrinks by ~gamma=0.9 each sweep (contraction mapping).", "N")
nt(ws_c, 3, "In-place converges faster: Warm immediately benefits from Cool update within the same sweep.", "N")

r = 5
ch = ["k",
      "V*(C) sync", "V*(W) sync", "|err C| sync", "|err W| sync", "|DV| sync", "% conv sync",
      "",
      "V*(C) inp", "V*(W) inp", "|err C| inp", "|err W| inp", "|DV| inp", "% conv inp"]
for j, h in enumerate(ch): ws_c.cell(r, 1+j, value=h)
hdr(ws_c, r, 1, 14)
ws_c.cell(r, 8).fill = PatternFill()

r = 6; ci = r
ws_c.cell(r,1,value=0); ws_c.cell(r,2,value=0); ws_c.cell(r,3,value=0)
fm(ws_c, r, 4, f"=ABS(B{r}-15.5)"); fm(ws_c, r, 5, f"=ABS(C{r}-14.5)")
ws_c.cell(r, 6, value="---")
fm(ws_c, r, 7, f"=1-MAX(D{r},E{r})/15.5")
ws_c.cell(r, 8, value="")
ws_c.cell(r,9,value=0); ws_c.cell(r,10,value=0)
fm(ws_c, r, 11, f"=ABS(I{r}-15.5)"); fm(ws_c, r, 12, f"=ABS(J{r}-14.5)")
ws_c.cell(r, 13, value="---")
fm(ws_c, r, 14, f"=1-MAX(K{r},L{r})/15.5")
dat(ws_c, r, 1, 14)

for k in range(1, ITERS + 1):
    r = ci + k; p = r - 1
    vr = vi_init + k; ir = ip_init + k
    ws_c.cell(r, 1, value=k)
    fm(ws_c, r, 2, f"='DP-VI'!D{vr}")
    fm(ws_c, r, 3, f"='DP-VI'!G{vr}")
    fm(ws_c, r, 4, f"=ABS(B{r}-15.5)")
    fm(ws_c, r, 5, f"=ABS(C{r}-14.5)")
    fm(ws_c, r, 6, f"=MAX(ABS(B{r}-B{p}),ABS(C{r}-C{p}))")
    fm(ws_c, r, 7, f"=1-MAX(D{r},E{r})/15.5")
    ws_c.cell(r, 8, value="")
    fm(ws_c, r, 9, f"='DP-VI'!D{ir}")
    fm(ws_c, r, 10, f"='DP-VI'!G{ir}")
    fm(ws_c, r, 11, f"=ABS(I{r}-15.5)")
    fm(ws_c, r, 12, f"=ABS(J{r}-14.5)")
    fm(ws_c, r, 13, f"=MAX(ABS(I{r}-I{p}),ABS(J{r}-J{p}))")
    fm(ws_c, r, 14, f"=1-MAX(K{r},L{r})/15.5")
    dat(ws_c, r, 1, 1); dat(ws_c, r, 8, 8)

cl = ci + ITERS

r = cl + 2; sec(ws_c, r, "Convergence Summary")
r += 1
for j, h in enumerate(["Metric", "Synchronous", "In-Place"]): ws_c.cell(r, 1+j, value=h)
hdr(ws_c, r, 1, 3); r += 1
for lbl, sf, ipf in [
    ("Final V*(Cool)", f"=B{cl}", f"=I{cl}"),
    ("Final V*(Warm)", f"=C{cl}", f"=J{cl}"),
    ("Analytical V*(Cool)", None, None),
    ("Analytical V*(Warm)", None, None),
    ("Final |error|", f"=MAX(D{cl},E{cl})", f"=MAX(K{cl},L{cl})"),
    ("Final % converged", f"=G{cl}", f"=N{cl}"),
    ("Final |DV|", f"=F{cl}", f"=M{cl}"),
    ("Contraction |DV_k|/|DV_{k-1}| at k=10", f"=F{ci+10}/F{ci+9}", f"=M{ci+10}/M{ci+9}"),
    ("Expected ratio (gamma)", None, None),
]:
    ws_c.cell(r, 1, value=lbl)
    if sf is None:
        if "Cool" in lbl: ws_c.cell(r, 2, value=15.5); ws_c.cell(r, 3, value=15.5)
        elif "Warm" in lbl: ws_c.cell(r, 2, value=14.5); ws_c.cell(r, 3, value=14.5)
        elif "gamma" in lbl: ws_c.cell(r, 2, value=0.9); ws_c.cell(r, 3, value=0.9)
    else:
        fm(ws_c, r, 2, sf); fm(ws_c, r, 3, ipf)
    dat(ws_c, r, 1, 3); r += 1

r += 1; sec(ws_c, r, "Convergence Speed (first k where |DV| < threshold)")
r += 1
for j, h in enumerate(["Threshold", "Sync k", "InPlace k", "InPlace faster by"]): ws_c.cell(r, 1+j, value=h)
hdr(ws_c, r, 1, 4); r += 1

sr = f"F{ci+1}:F{cl}"
ipr = f"M{ci+1}:M{cl}"

for th in [1.0, 0.1, 0.01, 0.001, 0.0001]:
    ws_c.cell(r, 1, value=th)
    # SUMPRODUCT finds first k where |DV| < threshold (no array entry needed)
    fm(ws_c, r, 2, f'=IFERROR(SUMPRODUCT(({sr}>={th})*1)+1,">{ITERS}")')
    fm(ws_c, r, 3, f'=IFERROR(SUMPRODUCT(({ipr}>={th})*1)+1,">{ITERS}")')
    fm(ws_c, r, 4, f'=IF(AND(ISNUMBER(B{r}),ISNUMBER(C{r})),B{r}-C{r},"n/a")')
    dat(ws_c, r, 1, 4); r += 1

aw(ws_c)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET: Comparisons — DP-Eval vs DP-VI vs Q-Learning
# ══════════════════════════════════════════════════════════════════════════════
ws_cmp = wb.create_sheet("Comparisons")
CW = "K"

ttl(ws_cmp, 1, "Method Comparisons: DP-Eval vs DP-VI vs Model-Free Methods", CW)

# ══════════════════════════════════════════════════════════════════════════════
# PART 1: DP-Eval vs DP-VI
# ══════════════════════════════════════════════════════════════════════════════
r = 3; sec(ws_cmp, r, "Part 1: DP-Eval (Policy Evaluation) vs DP-VI (Value Iteration)", "2F5496")
r += 1
nt(ws_cmp, r, "Both are Dynamic Programming methods that sweep over ALL states using the KNOWN model. The difference is the OPERATOR: expectation vs max.", CW)
r += 2

# Overview table
cmp1_hdrs = ["Aspect", "DP-Eval (Policy Evaluation)", "DP-VI (Value Iteration)"]
for j, h in enumerate(cmp1_hdrs): ws_cmp.cell(r, 1+j, value=h)
ws_cmp.merge_cells(f"B{r}:F{r}"); ws_cmp.merge_cells(f"G{r}:{CW}{r}")
hdr(ws_cmp, r, 1, 11); r += 1

cmp1_rows = [
    ("Question it answers",
     "How good is THIS specific policy pi?",
     "What is the BEST possible policy?"),
    ("Bellman equation used",
     "Expectation: V(s) = SUM P*[R + gamma*V(s')]  under FIXED pi",
     "Optimality: V(s) = MAX_a SUM P*[R + gamma*V(s')]  over ALL actions"),
    ("Key operator",
     "NONE (just follow the given policy)",
     "MAX over actions (find the best one)"),
    ("Input required",
     "A specific policy (e.g. 'always slow')",
     "No policy needed — finds optimal automatically"),
    ("Output",
     "V^pi(s) for that one policy",
     "V*(s), Q*(s,a), and pi*(s) — all optimal"),
    ("Columns in the sheet",
     "V_k(Cool), V_k(Warm), |DV|  (no Q columns)",
     "Q(s,Slow), Q(s,Fast), V*=MAX(Q), pi*=argmax(Q), |DV|"),
    ("Number of runs needed",
     "One per policy (4 policies = 4 blocks)",
     "One run finds everything"),
    ("Uses transition probs P?",
     "YES — embedded in the formula",
     "YES — embedded in the formula"),
    ("Convergence target",
     "Different for each policy (Cautious->10, Optimal->15.5)",
     "Always V*: V*(Cool)=15.5, V*(Warm)=14.5"),
]
for aspect, eval_desc, vi_desc in cmp1_rows:
    ws_cmp.cell(r, 1, value=aspect); ws_cmp.cell(r, 1).font = Font(bold=True)
    ws_cmp.cell(r, 2, value=eval_desc); ws_cmp.merge_cells(f"B{r}:F{r}")
    ws_cmp.cell(r, 2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws_cmp.cell(r, 7, value=vi_desc); ws_cmp.merge_cells(f"G{r}:{CW}{r}")
    ws_cmp.cell(r, 7).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    dat(ws_cmp, r, 1, 11); r += 1

# Formula comparison
r += 1; sec(ws_cmp, r, "Formula Comparison — Same (Cool, Slow) transition, different equations:")
r += 1
form_hdrs = ["", "Formula", "Explanation"]
for j, h in enumerate(form_hdrs): ws_cmp.cell(r, 1+j, value=h)
ws_cmp.merge_cells(f"B{r}:F{r}"); ws_cmp.merge_cells(f"G{r}:{CW}{r}")
hdr(ws_cmp, r, 1, 11); r += 1

formulas_cmp = [
    ("DP-Eval\n(Cautious: always Slow)",
     "V_{k+1}(Cool) = 0.5*(1+0.9*V_k(C)) + 0.5*(1+0.9*V_k(W))",
     "Policy says Slow, so we ONLY compute the Slow outcome. No choice involved — just evaluate what this policy gives us."),
    ("DP-Eval\n(Optimal: Fast@Cool)",
     "V_{k+1}(Cool) = 0.5*(2+0.9*V_k(C)) + 0.5*(2+0.9*V_k(W))",
     "Policy says Fast at Cool, so we compute the Fast outcome. Still no MAX — policy is fixed."),
    ("DP-VI",
     "V_{k+1}(Cool) = MAX( 0.5*(1+0.9*V_k(C))+0.5*(1+0.9*V_k(W)),  0.5*(2+0.9*V_k(C))+0.5*(2+0.9*V_k(W)) )",
     "Computes BOTH Slow and Fast Q-values, then takes MAX. This IS how it finds the optimal policy — it tries everything."),
]
for label, formula, expl in formulas_cmp:
    ws_cmp.cell(r, 1, value=label); ws_cmp.cell(r, 1).font = Font(bold=True, size=10)
    ws_cmp.cell(r, 1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws_cmp.cell(r, 2, value=formula); ws_cmp.merge_cells(f"B{r}:F{r}")
    ws_cmp.cell(r, 2).font = Font(name="Courier New", size=10)
    ws_cmp.cell(r, 2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws_cmp.cell(r, 7, value=expl); ws_cmp.merge_cells(f"G{r}:{CW}{r}")
    ws_cmp.cell(r, 7).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    dat(ws_cmp, r, 1, 11); r += 1

# How they relate
r += 1; sec(ws_cmp, r, "How DP-Eval and DP-VI relate (Policy Iteration uses BOTH):")
r += 1
pi_steps = [
    "1. START with any policy (e.g. Cautious: always Slow)",
    "2. EVALUATE it using DP-Eval -> get V^cautious(Cool)=10, V^cautious(Warm)=10",
    "3. IMPROVE: for each state, compute Q^cautious(s,a) for all actions, pick argmax -> new policy = Fast@Cool, Slow@Warm",
    "4. EVALUATE the new policy using DP-Eval -> get V^optimal(Cool)=15.5, V^optimal(Warm)=14.5",
    "5. IMPROVE: argmax Q^optimal -> same policy (Fast@Cool, Slow@Warm) -> CONVERGED!",
    "",
    "DP-VI is a shortcut: it combines evaluate+improve into one step by using MAX directly.",
    "Instead of evaluating a policy fully then improving, it improves greedily at EVERY sweep.",
]
for step in pi_steps:
    ws_cmp.merge_cells(f"A{r}:{CW}{r}")
    ws_cmp.cell(r, 1, value=step)
    if step.startswith("DP-VI"):
        ws_cmp.cell(r, 1).font = Font(bold=True, size=11, color="C00000")
    else:
        ws_cmp.cell(r, 1).font = Font(size=11)
    ws_cmp.cell(r, 1).alignment = Alignment(horizontal="left", vertical="center")
    r += 1

# Live formula proof: DP-Eval converged values
r += 1; sec(ws_cmp, r, "Live Proof: DP-Eval converged values vs DP-VI converged values")
r += 1
proof_hdrs = ["Value", "DP-Eval (Cautious)", "DP-Eval (Optimal)", "DP-VI", "Eval-Opt = VI?"]
for j, h in enumerate(proof_hdrs): ws_cmp.cell(r, 1+j, value=h)
hdr(ws_cmp, r, 1, 5); r += 1

ws_cmp.cell(r, 1, value="V(Cool)"); ws_cmp.cell(r, 1).font = Font(bold=True)
fm(ws_cmp, r, 2, f"='DP-Eval'!B{pe_caut}")
fm(ws_cmp, r, 3, f"='DP-Eval'!B{pe_opt}")
fm(ws_cmp, r, 4, f"='DP-VI'!D{vi_last}")
fm(ws_cmp, r, 5, f'=IF(ABS(C{r}-D{r})<0.001,"YES — same V*","NO")')
dat(ws_cmp, r, 1, 1); r += 1

ws_cmp.cell(r, 1, value="V(Warm)"); ws_cmp.cell(r, 1).font = Font(bold=True)
fm(ws_cmp, r, 2, f"='DP-Eval'!C{pe_caut}")
fm(ws_cmp, r, 3, f"='DP-Eval'!C{pe_opt}")
fm(ws_cmp, r, 4, f"='DP-VI'!G{vi_last}")
fm(ws_cmp, r, 5, f'=IF(ABS(C{r}-D{r})<0.001,"YES — same V*","NO")')
dat(ws_cmp, r, 1, 1); r += 1

r += 1
nt(ws_cmp, r, "DP-Eval(Cautious) = 10: the cautious policy is suboptimal.  DP-Eval(Optimal) = DP-VI = 15.5/14.5: evaluating the optimal policy gives V*.", CW)

# ══════════════════════════════════════════════════════════════════════════════
# PART 2: DP-VI (Model-Based) vs Q-Learning (Model-Free)
# ══════════════════════════════════════════════════════════════════════════════
r += 3; sec(ws_cmp, r, "Part 2: DP-VI (Model-Based) vs Q-Learning (Model-Free)", "C00000")
r += 1
nt(ws_cmp, r, "Both find Q* and pi*. The fundamental difference: DP-VI CALCULATES expectations using known P(s'|s,a). Q-Learning ESTIMATES them from samples.", CW)
r += 2

# Overview table
cmp2_hdrs = ["Aspect", "DP-VI (Model-Based)", "Q-Learning (Model-Free)"]
for j, h in enumerate(cmp2_hdrs): ws_cmp.cell(r, 1+j, value=h)
ws_cmp.merge_cells(f"B{r}:F{r}"); ws_cmp.merge_cells(f"G{r}:{CW}{r}")
hdr(ws_cmp, r, 1, 11); r += 1

cmp2_rows = [
    ("Knows P(s'|s,a)?",
     "YES — hardcoded in the formula as 0.5, 0.5, 1.0, etc.",
     "NO — never sees transition probabilities"),
    ("Knows R(s,a,s')?",
     "YES — hardcoded as +1, +2, -10",
     "NO — only observes the reward r after each transition"),
    ("How it sees next states",
     "Sums over ALL possible s' simultaneously",
     "Observes ONE random s' per step (whatever happened)"),
    ("Update formula",
     "Q(s,a) = SUM_{s'} P(s'|s,a)*[R + gamma*max Q(s',*)]",
     "Q(s,a) <- Q(s,a) + alpha*[r + gamma*max Q(s',*) - Q(s,a)]"),
    ("Learning rate alpha",
     "alpha = 1 (replace with exact target)",
     "alpha = 0.1 (inch toward noisy target)"),
    ("Coverage per step",
     "One SWEEP updates ALL (state, action) pairs",
     "One step updates ONE (state, action) pair"),
    ("Convergence speed",
     "Fast: ~50 sweeps (deterministic)",
     "Slow: needs many episodes (stochastic)"),
    ("When to use",
     "When you KNOW the model (game rules, physics equations)",
     "When you DON'T know the model (real world, complex simulations)"),
]
for aspect, dp_desc, ql_desc in cmp2_rows:
    ws_cmp.cell(r, 1, value=aspect); ws_cmp.cell(r, 1).font = Font(bold=True)
    ws_cmp.cell(r, 2, value=dp_desc); ws_cmp.merge_cells(f"B{r}:F{r}")
    ws_cmp.cell(r, 2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws_cmp.cell(r, 7, value=ql_desc); ws_cmp.merge_cells(f"G{r}:{CW}{r}")
    ws_cmp.cell(r, 7).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    dat(ws_cmp, r, 1, 11); r += 1

# The key formula comparison — where P appears
r += 1; sec(ws_cmp, r, "WHERE the transition probability P appears (and doesn't):", "C00000")
r += 1
nt(ws_cmp, r, "This is THE critical difference. Look at Q(Cool, Fast) computed both ways:", CW)
r += 2

ws_cmp.cell(r, 1, value="DP-VI formula:"); ws_cmp.cell(r, 1).font = Font(bold=True, size=12, color="2F5496")
ws_cmp.merge_cells(f"B{r}:{CW}{r}")
ws_cmp.cell(r, 2, value="Q(Cool,Fast) = 0.5 * (2 + 0.9*V(C))  +  0.5 * (2 + 0.9*V(W))")
ws_cmp.cell(r, 2).font = Font(name="Courier New", size=12)
ws_cmp.cell(r, 2).alignment = Alignment(horizontal="left", vertical="center")
dat(ws_cmp, r, 1, 11); r += 1

ws_cmp.merge_cells(f"B{r}:{CW}{r}")
ws_cmp.cell(r, 2, value="                 ^^^                     ^^^")
ws_cmp.cell(r, 2).font = Font(name="Courier New", size=12, color="C00000")
ws_cmp.cell(r, 2).alignment = Alignment(horizontal="left", vertical="center")
r += 1

ws_cmp.merge_cells(f"B{r}:{CW}{r}")
ws_cmp.cell(r, 2, value="            P(Cool|Cool,Fast)=0.5    P(Warm|Cool,Fast)=0.5")
ws_cmp.cell(r, 2).font = Font(name="Courier New", size=11, color="C00000")
ws_cmp.cell(r, 2).alignment = Alignment(horizontal="left", vertical="center")
r += 2

ws_cmp.cell(r, 1, value="Q-Learning step:"); ws_cmp.cell(r, 1).font = Font(bold=True, size=12, color="00B050")
ws_cmp.merge_cells(f"B{r}:{CW}{r}")
ws_cmp.cell(r, 2, value="Agent picks Fast in Cool -> lands in Warm (random), gets r=2")
ws_cmp.cell(r, 2).font = Font(name="Courier New", size=11)
ws_cmp.cell(r, 2).alignment = Alignment(horizontal="left", vertical="center")
dat(ws_cmp, r, 1, 11); r += 1

ws_cmp.merge_cells(f"B{r}:{CW}{r}")
ws_cmp.cell(r, 2, value="target = 2 + 0.9 * max(Q(Warm,Slow), Q(Warm,Fast))")
ws_cmp.cell(r, 2).font = Font(name="Courier New", size=11)
ws_cmp.cell(r, 2).alignment = Alignment(horizontal="left", vertical="center")
r += 1

ws_cmp.merge_cells(f"B{r}:{CW}{r}")
ws_cmp.cell(r, 2, value="Q(Cool,Fast) <- Q(Cool,Fast) + 0.1 * (target - Q(Cool,Fast))")
ws_cmp.cell(r, 2).font = Font(name="Courier New", size=11)
ws_cmp.cell(r, 2).alignment = Alignment(horizontal="left", vertical="center")
r += 1

ws_cmp.merge_cells(f"B{r}:{CW}{r}")
ws_cmp.cell(r, 2, value="   ^^^ NO 0.5 anywhere! Only saw ONE outcome (Warm). Next time might be Cool.")
ws_cmp.cell(r, 2).font = Font(name="Courier New", size=11, color="C00000")
ws_cmp.cell(r, 2).alignment = Alignment(horizontal="left", vertical="center")
r += 2

# How Q-Learning recovers the same answer
sec(ws_cmp, r, "How Q-Learning recovers P without knowing it:")
r += 1
recovery = [
    "Visit 1: Fast in Cool -> lands in Warm (r=2).   Q updated toward (2 + 0.9*max Q(Warm,.))",
    "Visit 2: Fast in Cool -> lands in Cool (r=2).   Q updated toward (2 + 0.9*max Q(Cool,.))",
    "Visit 3: Fast in Cool -> lands in Warm (r=2).   Q updated toward (2 + 0.9*max Q(Warm,.))",
    "Visit 4: Fast in Cool -> lands in Cool (r=2).   Q updated toward (2 + 0.9*max Q(Cool,.))",
    "...",
    "After many visits: ~50% landed in Cool, ~50% landed in Warm.",
    "The running average of targets CONVERGES to: 0.5*(2+0.9*V(C)) + 0.5*(2+0.9*V(W))",
    "...which is EXACTLY the DP-VI formula! The probabilities emerge from sampling frequency.",
]
for line in recovery:
    ws_cmp.merge_cells(f"A{r}:{CW}{r}")
    ws_cmp.cell(r, 1, value=line)
    if line.startswith("After") or line.startswith("The running") or line.startswith("...which"):
        ws_cmp.cell(r, 1).font = Font(bold=True, size=11, color="00B050")
    else:
        ws_cmp.cell(r, 1).font = Font(name="Courier New", size=10)
    ws_cmp.cell(r, 1).alignment = Alignment(horizontal="left", vertical="center")
    r += 1

# Live proof: DP-VI vs Q-Learning converged values
r += 1; sec(ws_cmp, r, "Live Proof: DP-VI vs Q-Learning converged Q-values")
r += 1
nt(ws_cmp, r, "Q-Learning has only 5 episodes (76 steps) so it hasn't fully converged yet. With more episodes it would approach DP-VI exactly.", CW)
r += 1

ql_qt_ref = 6 + len(all_steps) + 4  # ql_init=6, ql_last=6+len(all_steps), Q-table header+1 row
proof2_hdrs = ["Q(s,a)", "DP-VI (exact)", "Q-Learning (5 eps)", "Gap", "Converging?"]
for j, h in enumerate(proof2_hdrs): ws_cmp.cell(r, 1+j, value=h)
hdr(ws_cmp, r, 1, 5); r += 1

# Q(Cool,Slow): DP-VI sync last row col B vs QLearning final Q(C,S) col L
ws_cmp.cell(r, 1, value="Q(Cool, Slow)"); ws_cmp.cell(r, 1).font = Font(bold=True)
fm(ws_cmp, r, 2, f"='DP-VI'!B{vi_last}")
fm(ws_cmp, r, 3, f"='QLearning'!B{ql_qt_ref}")
fm(ws_cmp, r, 4, f"=ABS(B{r}-C{r})")
fm(ws_cmp, r, 5, f'=IF(D{r}<1,"Close","Far")')
dat(ws_cmp, r, 1, 1); r += 1

ws_cmp.cell(r, 1, value="Q(Cool, Fast)"); ws_cmp.cell(r, 1).font = Font(bold=True)
fm(ws_cmp, r, 2, f"='DP-VI'!C{vi_last}")
fm(ws_cmp, r, 3, f"='QLearning'!C{ql_qt_ref}")
fm(ws_cmp, r, 4, f"=ABS(B{r}-C{r})")
fm(ws_cmp, r, 5, f'=IF(D{r}<1,"Close","Far")')
dat(ws_cmp, r, 1, 1); r += 1

ws_cmp.cell(r, 1, value="Q(Warm, Slow)"); ws_cmp.cell(r, 1).font = Font(bold=True)
fm(ws_cmp, r, 2, f"='DP-VI'!E{vi_last}")
fm(ws_cmp, r, 3, f"='QLearning'!B{ql_qt_ref+1}")
fm(ws_cmp, r, 4, f"=ABS(B{r}-C{r})")
fm(ws_cmp, r, 5, f'=IF(D{r}<1,"Close","Far")')
dat(ws_cmp, r, 1, 1); r += 1

ws_cmp.cell(r, 1, value="Q(Warm, Fast)"); ws_cmp.cell(r, 1).font = Font(bold=True)
fm(ws_cmp, r, 2, f"='DP-VI'!F{vi_last}")
fm(ws_cmp, r, 3, f"='QLearning'!C{ql_qt_ref+1}")
fm(ws_cmp, r, 4, f"=ABS(B{r}-C{r})")
fm(ws_cmp, r, 5, f'=IF(D{r}<1,"Close","Far")')
dat(ws_cmp, r, 1, 1); r += 1

# ══════════════════════════════════════════════════════════════════════════════
# PART 3: Complete Method Family Comparison
# ══════════════════════════════════════════════════════════════════════════════
r += 2; sec(ws_cmp, r, "Part 3: Complete Comparison — All Five Methods in This Workbook", "375623")
r += 1
nt(ws_cmp, r, "Every method converges to the same V* and pi*. They differ in what information they need and how they use it.", CW)
r += 2

all_hdrs = ["", "DP-Eval", "DP-VI", "MC", "SARSA", "Q-Learning"]
for j, h in enumerate(all_hdrs): ws_cmp.cell(r, 1+j, value=h)
hdr(ws_cmp, r, 1, 6); r += 1

all_rows = [
    ("Type",           "Model-based",    "Model-based",    "Model-free",     "Model-free",     "Model-free"),
    ("Finds V* or V^pi?", "V^pi only",  "V* (optimal)",   "V^pi estimate",  "Q^pi estimate",  "Q* estimate"),
    ("Uses P(s'|s,a)?","YES",            "YES",            "NO",             "NO",             "NO"),
    ("Uses R(s,a,s')?","YES",            "YES",            "NO (observes r)","NO (observes r)","NO (observes r)"),
    ("Policy needed?", "YES (fixed pi)", "NO (finds pi*)", "YES (behavior)", "YES (epsilon-greedy)","NO (off-policy)"),
    ("Update unit",    "Full sweep",     "Full sweep",     "Full episode",   "Single step",    "Single step"),
    ("alpha",          "1 (exact)",      "1 (exact)",      "1/N (average)",  "0.1 (small)",    "0.1 (small)"),
    ("Bellman eq.",    "Expectation",    "Optimality",     "None (returns)", "Expectation",    "Optimality"),
    ("On/Off-policy",  "N/A",           "N/A",            "On-policy",      "On-policy",      "Off-policy"),
    ("Sheet",          "DP-Eval",        "DP-VI",          "MC-Episodes/VQ", "SARSA",          "QLearning"),
]
for row_data in all_rows:
    for j, v in enumerate(row_data):
        ws_cmp.cell(r, 1+j, value=v)
        if j == 0:
            ws_cmp.cell(r, 1).font = Font(bold=True)
    dat(ws_cmp, r, 1, 6); r += 1

r += 1
summary_lines = [
    "KEY INSIGHT: The 0.5 in DP-VI's formula IS the transition probability. Q-Learning doesn't have it —",
    "it recovers the same weighted average by visiting each outcome enough times.",
    "DP-VI = analytical weighted sum.  Q-Learning = empirical running average.  Same destination, different roads.",
]
for line in summary_lines:
    ws_cmp.merge_cells(f"A{r}:{CW}{r}")
    ws_cmp.cell(r, 1, value=line)
    ws_cmp.cell(r, 1).font = Font(bold=True, size=11, color="C00000") if "KEY" in line else Font(italic=True, size=11)
    ws_cmp.cell(r, 1).alignment = Alignment(horizontal="left", vertical="center")
    r += 1

aw(ws_cmp)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET: Analytical Solution — Deriving V*, Q* from Bellman Equations
# ══════════════════════════════════════════════════════════════════════════════
ws_a = wb.create_sheet("Analytical")
ttl(ws_a, 1, "Analytical Solution: Solving the Bellman Optimality Equation for V* and Q*", "G")

# ── Step 1: The Bellman Optimality Equation ──
r = 3; sec(ws_a, r, "Step 1: The Bellman Optimality Equation")
r += 1
nt(ws_a, r, "V*(s) = max_a  SUM_{s'} P(s'|s,a) * [ R(s,a,s') + gamma * V*(s') ]", "G"); r += 1
nt(ws_a, r, "This is a SYSTEM of simultaneous equations. V*(s) on the left and V*(s') on the right are the SAME solution.", "G"); r += 1
nt(ws_a, r, "It is NOT an iterative update — it describes the fixed point that DP converges TO.", "G"); r += 1

# ── Step 2: Write out Q*(s,a) for every (s,a) pair ──
r += 1; sec(ws_a, r, "Step 2: Expand Q*(s,a) = SUM P(s'|s,a) * [R + gamma*V*(s')] for each (s,a)")
r += 1
for j, h in enumerate(["(s, a)", "Expansion", "Simplified"]): ws_a.cell(r, 1+j, value=h)
hdr(ws_a, r, 1, 3); r += 1

qa_rows = {}
expansions = [
    ("(Cool, Slow)",
     "P(C|C,S)*[R(C,S,C)+g*V*(C)] + P(W|C,S)*[R(C,S,W)+g*V*(W)]",
     "0.5*(1+0.9*V*(C)) + 0.5*(1+0.9*V*(W))"),
    ("(Cool, Fast)",
     "P(C|C,F)*[R(C,F,C)+g*V*(C)] + P(W|C,F)*[R(C,F,W)+g*V*(W)]",
     "0.5*(2+0.9*V*(C)) + 0.5*(2+0.9*V*(W))"),
    ("(Warm, Slow)",
     "P(C|W,S)*[R(W,S,C)+g*V*(C)] + P(W|W,S)*[R(W,S,W)+g*V*(W)]",
     "0.5*(1+0.9*V*(C)) + 0.5*(1+0.9*V*(W))"),
    ("(Warm, Fast)",
     "P(OH|W,F)*[R(W,F,OH)+g*V*(OH)]",
     "1*(-10+0.9*0) = -10"),
]
for sa, exp, simp in expansions:
    ws_a.cell(r, 1, value=sa); ws_a.cell(r, 1).font = Font(bold=True)
    ws_a.cell(r, 2, value=exp); ws_a.cell(r, 3, value=simp)
    dat(ws_a, r, 1, 3)
    qa_rows[sa] = r; r += 1

# ── Step 3: Circular Dependency — Why it's a SYSTEM, not simple recursion ──
r += 1; sec(ws_a, r, "Step 3: Why This Is a SYSTEM of Equations (Circular Dependency)")
r += 1
nt(ws_a, r, "In TRUE recursion, you start from a base case (leaf) and work backward — each unknown depends only on ALREADY-KNOWN values.", "G"); r += 1
nt(ws_a, r, "Here, the Bellman equation is CIRCULAR: V*(Cool) needs V*(Warm), and V*(Warm) needs V*(Cool). Neither can be solved first.", "G"); r += 1
r += 1

# Dependency chain table
sec(ws_a, r, "Trace the dependency chain:"); r += 1
dep_hdrs = ["#", "To compute...", "We need...", "Which requires...", "Problem"]
for j, h in enumerate(dep_hdrs): ws_a.cell(r, 1+j, value=h)
hdr(ws_a, r, 1, 5); r += 1

dep_rows = [
    ("1", "V*(Cool) = max Q*(Cool,a)",
     "Q*(Cool,Fast) = 0.5*(2+0.9*V*(C)) + 0.5*(2+0.9*V*(W))",
     "V*(Cool) AND V*(Warm)",
     "V*(Warm) is UNKNOWN"),
    ("2", "V*(Warm) = max Q*(Warm,a)",
     "Q*(Warm,Slow) = 0.5*(1+0.9*V*(C)) + 0.5*(1+0.9*V*(W))",
     "V*(Cool) AND V*(Warm)",
     "V*(Cool) is UNKNOWN"),
    ("3", "Attempt: solve V*(Cool) first?",
     "Stuck — need V*(Warm) which needs V*(Cool)",
     "CIRCULAR: Cool -> Warm -> Cool -> ...",
     "No ordering exists!"),
    ("4", "Attempt: solve V*(Warm) first?",
     "Stuck — need V*(Cool) which needs V*(Warm)",
     "CIRCULAR: Warm -> Cool -> Warm -> ...",
     "Same deadlock!"),
]
for num, comp, need, req, prob in dep_rows:
    ws_a.cell(r, 1, value=num); ws_a.cell(r, 1).font = Font(bold=True)
    ws_a.cell(r, 2, value=comp)
    ws_a.cell(r, 3, value=need)
    ws_a.cell(r, 4, value=req)
    ws_a.cell(r, 5, value=prob); ws_a.cell(r, 5).font = Font(bold=True, color="C00000")
    dat(ws_a, r, 1, 5); r += 1

r += 1
sec(ws_a, r, "The ROOT CAUSE: Cycles in the transition graph", "C00000"); r += 1
nt(ws_a, r, "Cool --[Fast/Slow]--> Cool or Warm.    Warm --[Slow]--> Cool or Warm.    This creates a CYCLE: Cool <-> Warm.", "G"); r += 1
nt(ws_a, r, "Cycles mean mutual dependency. The value of being in Cool depends on what happens in Warm, and vice versa.", "G"); r += 1
r += 1

# Contrast with what WOULD be simple recursion
sec(ws_a, r, "Contrast: When IS it simple recursion? (no cycles)"); r += 1
contrast_hdrs = ["Scenario", "Dependency", "Solvable by recursion?"]
for j, h in enumerate(contrast_hdrs): ws_a.cell(r, 1+j, value=h)
hdr(ws_a, r, 1, 3); r += 1

contrasts = [
    ("Q*(Warm, Fast) = -10",
     "Only depends on V*(OH) = 0 (terminal, KNOWN)",
     "YES — base case, no unknowns"),
    ("A -> B -> C -> Terminal (no cycles)",
     "V*(C) needs terminal only, V*(B) needs V*(C), V*(A) needs V*(B)",
     "YES — solve backward from terminal"),
    ("Cool <-> Warm (our problem!)",
     "V*(Cool) needs V*(Warm), V*(Warm) needs V*(Cool)",
     "NO — must solve simultaneously"),
    ("N states all connected to each other",
     "Every V*(s) depends on every other V*(s')",
     "NO — N equations in N unknowns"),
]
for scen, dep, solv in contrasts:
    ws_a.cell(r, 1, value=scen); ws_a.cell(r, 1).font = Font(bold=True)
    ws_a.cell(r, 2, value=dep)
    ws_a.cell(r, 3, value=solv)
    dat(ws_a, r, 1, 3); r += 1

r += 1
sec(ws_a, r, "Three ways to handle the circular dependency:"); r += 1
solutions = [
    "1. ALGEBRA (this sheet): Introduce variable x, substitute, solve the resulting equation. Works for small MDPs.",
    "2. LINEAR ALGEBRA: Write as matrix equation V = R + gamma*P*V, solve (I - gamma*P)*V = R. Works for any size.",
    "3. ITERATIVE (DP): Start with V=0, repeatedly apply Bellman update. Converges to the fixed point (see DP-VI sheet).",
]
for sol in solutions:
    ws_a.cell(r, 1, value=sol)
    ws_a.merge_cells(f"A{r}:G{r}")
    ws_a.cell(r, 1).font = Font(italic=True, size=10)
    r += 1

r += 1
nt(ws_a, r, "SUMMARY: It's called a 'system' because V*(Cool) and V*(Warm) are coupled unknowns that must be solved TOGETHER, not one after the other.", "G"); r += 1

# ── Step 4: Observe key relationships ──
r += 1; sec(ws_a, r, "Step 4: Key Observations (before solving)")
r += 1
observations = [
    "Q*(Cool,Slow) = Q*(Warm,Slow) — same formula because both have 0.5/0.5 to Cool/Warm with R=1",
    "Q*(Cool,Fast) = Q*(Cool,Slow) + 1 — Fast just adds +1 more reward per transition (2 vs 1)",
    "Q*(Warm,Fast) = -10 — no unknowns, Warm+Fast always goes to terminal Overheated",
    "Since Q*(Cool,Fast) > Q*(Cool,Slow) always: V*(Cool) = Q*(Cool,Fast) — optimal action at Cool is Fast",
    "Since Q*(Warm,Slow) >> Q*(Warm,Fast)=-10: V*(Warm) = Q*(Warm,Slow) — optimal action at Warm is Slow",
]
for obs in observations:
    ws_a.cell(r, 1, value=obs)
    ws_a.merge_cells(f"A{r}:G{r}")
    ws_a.cell(r, 1).font = Font(italic=True, size=10)
    r += 1

# ── Step 5: Set up simultaneous equations ──
r += 1; sec(ws_a, r, "Step 5: Set up the simultaneous equations")
r += 1
nt(ws_a, r, "Let x = Q*(Cool,Slow) = Q*(Warm,Slow).  Then V*(Warm) = x, V*(Cool) = x+1.", "G"); r += 1
r += 1
eqs = [
    ("Define:", "x = V*(Warm) = Q*(Warm,Slow)"),
    ("Then:", "V*(Cool) = Q*(Cool,Fast) = x + 1"),
    ("", ""),
    ("Substitute into Q*(Warm,Slow):", "x = 0.5*(1 + 0.9*(x+1)) + 0.5*(1 + 0.9*x)"),
]
for lbl, eq in eqs:
    ws_a.cell(r, 1, value=lbl); ws_a.cell(r, 1).font = Font(bold=True)
    ws_a.cell(r, 2, value=eq); ws_a.merge_cells(f"B{r}:G{r}")
    dat(ws_a, r, 1, 1); r += 1

# ── Step 6: Solve algebraically (each step is a formula) ──
r += 1; sec(ws_a, r, "Step 6: Solve step by step (every cell is a formula verifying each algebraic step)")
r += 1
for j, h in enumerate(["Step", "Equation", "LHS", "RHS", "LHS=RHS?"]): ws_a.cell(r, 1+j, value=h)
hdr(ws_a, r, 1, 5); r += 1

# Write algebra steps first (text only), then place x, then go back and fill formulas
step_texts = [
    ("Start",        "x = 0.5*(1+0.9*(x+1)) + 0.5*(1+0.9*x)"),
    ("Expand 1st",   "0.5*(1 + 0.9x + 0.9)"),
    ("Expand 2nd",   "0.5*(1 + 0.9x)"),
    ("Sum both",     "0.5+0.45x+0.45 + 0.5+0.45x"),
    ("Collect",      "1.45 + 0.9x"),
    ("Rearrange",    "x - 0.9x = 1.45"),
    ("Factor",       "0.1x = 1.45"),
    ("Solve",        "x = 1.45 / 0.1 = 14.5"),
]
step_start = r
for step_name, eq_text in step_texts:
    ws_a.cell(r, 1, value=step_name); ws_a.cell(r, 1).font = Font(bold=True)
    ws_a.cell(r, 2, value=eq_text); dat(ws_a, r, 1, 2); r += 1

# ── Step 7: The solution ── place x FIRST so formulas can reference it
r += 1; sec(ws_a, r, "Step 7: The Solution")
r += 1
x_cell_row = r  # x = V*(Warm) = 14.5 lives here

# Now back-fill verification formulas into the step rows (cols C=LHS, D=RHS, E=match)
X = f"B{x_cell_row}"
step_formulas = [
    None,  # "Start" — no verification
    (f"=0.5*(1+0.9*{X}+0.9)",       f"=0.5*(1+0.9*({X}+1))"),
    (f"=0.5*(1+0.9*{X})",           f"=0.5*(1+0.9*{X})"),
    (f"=0.5+0.45*{X}+0.45+0.5+0.45*{X}", f"=0.5*(1+0.9*({X}+1))+0.5*(1+0.9*{X})"),
    (f"=1.45+0.9*{X}",              f"=0.5*(1+0.9*({X}+1))+0.5*(1+0.9*{X})"),
    (f"={X}-0.9*{X}",               "=1.45"),
    (f"=0.1*{X}",                   "=1.45"),
    (f"={X}",                        "=1.45/0.1"),
]
for i, sf in enumerate(step_formulas):
    sr = step_start + i
    if sf:
        fm(ws_a, sr, 3, sf[0])
        fm(ws_a, sr, 4, sf[1])
        fm(ws_a, sr, 5, f'=IF(ABS(C{sr}-D{sr})<0.0001,"= YES","= NO")')

ws_a.cell(r, 1, value="x = V*(Warm) ="); ws_a.cell(r, 1).font = Font(bold=True, size=12)
ws_a.cell(r, 2, value=14.5); ws_a.cell(r, 2).font = Font(bold=True, size=12, color="C00000")
ws_a.cell(r, 2).fill = QF
dat(ws_a, r, 1, 2); r += 1

ws_a.cell(r, 1, value="V*(Cool) = x+1 ="); ws_a.cell(r, 1).font = Font(bold=True, size=12)
fm(ws_a, r, 2, f"=B{x_cell_row}+1"); ws_a.cell(r, 2).font = Font(bold=True, size=12, color="C00000")
ws_a.cell(r, 2).fill = QF
dat(ws_a, r, 1, 1); r += 1

ws_a.cell(r, 1, value="V*(Overheated) ="); ws_a.cell(r, 1).font = Font(bold=True, size=12)
ws_a.cell(r, 2, value=0); ws_a.cell(r, 2).font = Font(bold=True, size=12)
ws_a.cell(r, 2).fill = TF
dat(ws_a, r, 1, 2); r += 1

# ── Step 8: Derive all Q* from V* ──
r += 1; sec(ws_a, r, "Step 8: Compute all Q*(s,a) from V* (formula verification)")
r += 1
for j, h in enumerate(["(s, a)", "Q*(s,a) formula", "Q*(s,a) value", "Verify"]): ws_a.cell(r, 1+j, value=h)
hdr(ws_a, r, 1, 4); r += 1

vc_ref = f"B{x_cell_row+1}"  # V*(Cool) cell
vw_ref = f"B{x_cell_row}"    # V*(Warm) cell

q_analytical = {}
q_formulas = [
    ("(Cool, Slow)",  f"=0.5*(1+0.9*{vc_ref})+0.5*(1+0.9*{vw_ref})", 14.5),
    ("(Cool, Fast)",  f"=0.5*(2+0.9*{vc_ref})+0.5*(2+0.9*{vw_ref})", 15.5),
    ("(Warm, Slow)",  f"=0.5*(1+0.9*{vc_ref})+0.5*(1+0.9*{vw_ref})", 14.5),
    ("(Warm, Fast)",  "=1*(-10+0.9*0)", -10.0),
]
for sa, qf, expected in q_formulas:
    ws_a.cell(r, 1, value=sa); ws_a.cell(r, 1).font = Font(bold=True)
    fm(ws_a, r, 2, qf)
    ws_a.cell(r, 3, value=expected)
    fm(ws_a, r, 4, f'=IF(ABS(B{r}-C{r})<0.0001,"= YES","= NO")')
    dat(ws_a, r, 1, 1)
    q_analytical[sa] = r; r += 1

# ── Step 9: Verify V* = max_a Q* ──
r += 1; sec(ws_a, r, "Step 9: Verify V*(s) = max_a Q*(s,a)")
r += 1
for j, h in enumerate(["State", "max_a Q*(s,a)", "V*(s)", "Match?", "Optimal Action"]): ws_a.cell(r, 1+j, value=h)
hdr(ws_a, r, 1, 5); r += 1

for st, qs_row, qf_row, v_ref in [
    ("Cool", q_analytical["(Cool, Slow)"], q_analytical["(Cool, Fast)"], vc_ref),
    ("Warm", q_analytical["(Warm, Slow)"], q_analytical["(Warm, Fast)"], vw_ref),
]:
    ws_a.cell(r, 1, value=st); ws_a.cell(r, 1).font = Font(bold=True)
    fm(ws_a, r, 2, f"=MAX(B{qs_row},B{qf_row})")
    fm(ws_a, r, 3, f"={v_ref}")
    fm(ws_a, r, 4, f'=IF(ABS(B{r}-C{r})<0.0001,"= YES","= NO")')
    fm(ws_a, r, 5, f'=IF(B{qf_row}>B{qs_row},"Fast","Slow")')
    dat(ws_a, r, 1, 1); r += 1

# ── Step 10: Verify Bellman equation holds ──
r += 1; sec(ws_a, r, "Step 10: Verify Bellman equation holds — plug V* back in")
r += 1
nt(ws_a, r, "If V* is correct, then V*(s) = max_a SUM P(s'|s,a)*[R+g*V*(s')] must be satisfied exactly.", "G"); r += 1
r += 1
for j, h in enumerate(["State", "LHS: V*(s)", "RHS: max_a SUM...", "LHS = RHS?"]): ws_a.cell(r, 1+j, value=h)
hdr(ws_a, r, 1, 4); r += 1

ws_a.cell(r, 1, value="Cool"); ws_a.cell(r, 1).font = Font(bold=True)
fm(ws_a, r, 2, f"={vc_ref}")
fm(ws_a, r, 3, f"=MAX(0.5*(1+0.9*{vc_ref})+0.5*(1+0.9*{vw_ref}), 0.5*(2+0.9*{vc_ref})+0.5*(2+0.9*{vw_ref}))")
fm(ws_a, r, 4, f'=IF(ABS(B{r}-C{r})<0.0001,"= YES","= NO")')
dat(ws_a, r, 1, 1); r += 1

ws_a.cell(r, 1, value="Warm"); ws_a.cell(r, 1).font = Font(bold=True)
fm(ws_a, r, 2, f"={vw_ref}")
fm(ws_a, r, 3, f"=MAX(0.5*(1+0.9*{vc_ref})+0.5*(1+0.9*{vw_ref}), 1*(-10+0.9*0))")
fm(ws_a, r, 4, f'=IF(ABS(B{r}-C{r})<0.0001,"= YES","= NO")')
dat(ws_a, r, 1, 1); r += 1

ws_a.cell(r, 1, value="Overheated"); ws_a.cell(r, 1).font = Font(bold=True)
ws_a.cell(r, 2, value=0); ws_a.cell(r, 3, value=0)
fm(ws_a, r, 4, f'=IF(ABS(B{r}-C{r})<0.0001,"= YES","= NO")')
dat(ws_a, r, 1, 1)
for c in range(1, 5): ws_a.cell(r, c).fill = TF
r += 1

# ── Step 11: Similarly solve V^pi for Cautious policy (always slow) ──
r += 1; sec(ws_a, r, "Step 11: Analytical V^cautious (always slow) for comparison")
r += 1
nt(ws_a, r, "Under always-slow: V(C) = V(W) = 0.5*(1+0.9*V(C)) + 0.5*(1+0.9*V(W)).  Let y = V(C) = V(W).", "G"); r += 1
r += 1
caut_steps = [
    ("Equation:", "y = 0.5*(1+0.9*y) + 0.5*(1+0.9*y)"),
    ("Simplify:", "y = 1 + 0.9*y"),
    ("Rearrange:", "y - 0.9*y = 1"),
    ("Factor:", "0.1*y = 1"),
    ("Solve:", "y = 10"),
]
for lbl, eq in caut_steps:
    ws_a.cell(r, 1, value=lbl); ws_a.cell(r, 1).font = Font(bold=True)
    ws_a.cell(r, 2, value=eq); ws_a.merge_cells(f"B{r}:E{r}")
    dat(ws_a, r, 1, 1); r += 1

r += 1
ws_a.cell(r, 1, value="V^cautious(Cool) = V^cautious(Warm) =")
ws_a.cell(r, 1).font = Font(bold=True, size=12)
ws_a.cell(r, 2, value=10); ws_a.cell(r, 2).font = Font(bold=True, size=12, color="2F5496")
ws_a.cell(r, 2).fill = QF
dat(ws_a, r, 1, 2)
caut_y_row = r; r += 1

# Verify
ws_a.cell(r, 1, value="Verify:"); ws_a.cell(r, 1).font = Font(bold=True)
fm(ws_a, r, 2, f"=0.5*(1+0.9*B{caut_y_row})+0.5*(1+0.9*B{caut_y_row})")
fm(ws_a, r, 3, f'=IF(ABS(B{r}-B{caut_y_row})<0.0001,"= YES","= NO")')
dat(ws_a, r, 1, 3); r += 1

# ── Summary comparison ──
r += 1; sec(ws_a, r, "Summary: Analytical vs DP-converged values")
r += 1
for j, h in enumerate(["Value", "Analytical", "DP (50 sweeps)", "Error"]): ws_a.cell(r, 1+j, value=h)
hdr(ws_a, r, 1, 4); r += 1

ws_a.cell(r, 1, value="V*(Cool)"); ws_a.cell(r, 1).font = Font(bold=True)
fm(ws_a, r, 2, f"={vc_ref}"); fm(ws_a, r, 3, f"='DP-VI'!D{vi_last}"); fm(ws_a, r, 4, f"=ABS(B{r}-C{r})")
dat(ws_a, r, 1, 1); r += 1
ws_a.cell(r, 1, value="V*(Warm)"); ws_a.cell(r, 1).font = Font(bold=True)
fm(ws_a, r, 2, f"={vw_ref}"); fm(ws_a, r, 3, f"='DP-VI'!G{vi_last}"); fm(ws_a, r, 4, f"=ABS(B{r}-C{r})")
dat(ws_a, r, 1, 1); r += 1
ws_a.cell(r, 1, value="V^caut(Cool)"); ws_a.cell(r, 1).font = Font(bold=True)
fm(ws_a, r, 2, f"=B{caut_y_row}"); fm(ws_a, r, 3, f"='DP-Eval'!B{pe_caut}"); fm(ws_a, r, 4, f"=ABS(B{r}-C{r})")
dat(ws_a, r, 1, 1); r += 1
ws_a.cell(r, 1, value="V^caut(Warm)"); ws_a.cell(r, 1).font = Font(bold=True)
fm(ws_a, r, 2, f"=B{caut_y_row}"); fm(ws_a, r, 3, f"='DP-Eval'!C{pe_caut}"); fm(ws_a, r, 4, f"=ABS(B{r}-C{r})")
dat(ws_a, r, 1, 1)

aw(ws_a)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4: V, V*, Q, Q* Summary + Policy Iteration Demo
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Summary")
ttl(ws, 1, "V, V*, Q, Q* — Complete Summary + Policy Iteration", "F")
nt(ws, 2, "All values are formulas referencing DP-Eval / DP-VI converged results, or computing Q from V via Bellman.", "F")

# ── Section 1: V^π and V* ──
r = 4; sec(ws, r, "Section 1: State Value Functions")
r = 5
for j, h in enumerate(["State", "V^cautious ✱", "V^optimal ✱", "V^aggressive ✱", "V^worst ✱", "V* ✱"]):
    ws.cell(r, 1+j, value=h)
hdr(ws, r, 1, 6)
v_row = {}
for i, (st, t) in enumerate([("Cool", False), ("Warm", False), ("Overheated", True)]):
    r = 6 + i; v_row[st] = r
    ws.cell(r, 1, value=st); ws.cell(r, 1).font = Font(bold=True)
    if t:
        for c in range(1, 7): ws.cell(r, c, value=0 if c > 1 else st); ws.cell(r, c).fill = TF
        dat(ws, r, 1, 6)
    else:
        col_map = {"Cool": "B", "Warm": "C"}
        fm(ws, r, 2, f"='DP-Eval'!{col_map[st]}{pe_caut}")
        fm(ws, r, 3, f"='DP-Eval'!{col_map[st]}{pe_opt}")
        fm(ws, r, 4, f"='DP-Eval'!{col_map[st]}{pe_agg}")
        fm(ws, r, 5, f"='DP-Eval'!{col_map[st]}{pe_worst}")
        vi_col = {"Cool": "D", "Warm": "G"}
        fm(ws, r, 6, f"='DP-VI'!{vi_col[st]}{vi_last}")
        dat(ws, r, 1, 1)

# V^π references for Q computation
# B6=V^caut(C), B7=V^caut(W), C6=V^opt(C), C7=V^opt(W), etc.
vc_c, vc_w = "B6", "B7"     # V^cautious
vo_c, vo_w = "C6", "C7"     # V^optimal
va_c, va_w = "D6", "D7"     # V^aggressive
vw_c, vw_w = "E6", "E7"     # V^worst
vs_c, vs_w = "F6", "F7"     # V*

# ── Section 2: Q^π for each policy ──
def write_q_section(ws, start_r, title_text, color, v_c, v_w, label):
    r = start_r; sec(ws, r, title_text, color)
    r += 1
    nt(ws, r, f"Q(s,a) = Σ P(s'|s,a)·[R(s,a,s') + γ·{label}(s')]", "D")
    r += 1
    for j, h in enumerate(["(s, a)", f"Q^{label}(s,a) ✱"]): ws.cell(r, 1+j, value=h)
    hdr(ws, r, 1, 2); r += 1
    rows = {}
    for sa, f in [
        ("(Cool, Slow)",  f"=0.5*(1+0.9*{v_c})+0.5*(1+0.9*{v_w})"),
        ("(Cool, Fast)",  f"=0.5*(2+0.9*{v_c})+0.5*(2+0.9*{v_w})"),
        ("(Warm, Slow)",  f"=0.5*(1+0.9*{v_c})+0.5*(1+0.9*{v_w})"),
        ("(Warm, Fast)",  "=1*(-10+0.9*0)"),
    ]:
        ws.cell(r, 1, value=sa); fm(ws, r, 2, f); dat(ws, r, 1, 1); rows[sa] = r; r += 1
    return r, rows

r_next, q_caut_rows = write_q_section(ws, 10, "Section 2a: Q^cautious(s,a)", "2F5496", vc_c, vc_w, "cautious")
r_next, q_opt_rows = write_q_section(ws, r_next + 1, "Section 2b: Q^optimal(s,a)", "00B050", vo_c, vo_w, "optimal")
r_next, q_agg_rows = write_q_section(ws, r_next + 1, "Section 2c: Q^aggressive(s,a)", "C00000", va_c, va_w, "aggressive")
r_next, q_worst_rows = write_q_section(ws, r_next + 1, "Section 2d: Q^worst(s,a)", "7030A0", vw_c, vw_w, "worst")
r_next, q_star_rows = write_q_section(ws, r_next + 1, "Section 2e: Q*(s,a) from V*", "375623", vs_c, vs_w, "*")

# ── Section 3: V*(s) = max Q*(s,a) verification ──
r = r_next + 1; sec(ws, r, "Section 3: Verify V*(s) = max_a Q*(s,a)")
r += 1
for j, h in enumerate(["State", "max_a Q*(s,a) ✱", "V*(s) ✱", "Match? ✱"]): ws.cell(r, 1+j, value=h)
hdr(ws, r, 1, 4); r += 1
vstar_verify_start = r
for st, q_s, q_f in [("Cool", q_star_rows["(Cool, Slow)"], q_star_rows["(Cool, Fast)"]),
                       ("Warm", q_star_rows["(Warm, Slow)"], q_star_rows["(Warm, Fast)"])]:
    ws.cell(r, 1, value=st); ws.cell(r, 1).font = Font(bold=True)
    fm(ws, r, 2, f"=MAX(B{q_s},B{q_f})")
    fm(ws, r, 3, f"=F{v_row[st]}")
    fm(ws, r, 4, f'=IF(ABS(B{r}-C{r})<0.01,"✓ Yes","✗ No")')
    dat(ws, r, 1, 1); r += 1

# ── Section 4: Policy Iteration Demo ──
r += 1; sec(ws, r, "Section 4: Policy Iteration Demonstration")
r += 1; nt(ws, r, "Step 1: Evaluate Cautious → Q^caut.  Step 2: Improve π(s)=argmax Q^caut → Optimal.  Step 3: Evaluate Optimal → Q^opt.  Step 4: argmax → same Optimal = converged.", "F")
r += 1
for j, h in enumerate(["State", "argmax Q^caut ✱", "argmax Q^opt ✱", "Converged? ✱"]): ws.cell(r, 1+j, value=h)
hdr(ws, r, 1, 4); r += 1
for st, (qs, qf), (qs2, qf2) in [
    ("Cool", (q_caut_rows["(Cool, Slow)"], q_caut_rows["(Cool, Fast)"]),
             (q_opt_rows["(Cool, Slow)"], q_opt_rows["(Cool, Fast)"])),
    ("Warm", (q_caut_rows["(Warm, Slow)"], q_caut_rows["(Warm, Fast)"]),
             (q_opt_rows["(Warm, Slow)"], q_opt_rows["(Warm, Fast)"])),
]:
    ws.cell(r, 1, value=st); ws.cell(r, 1).font = Font(bold=True)
    fm(ws, r, 2, f'=IF(B{qf}>=B{qs},"Fast","Slow")')
    fm(ws, r, 3, f'=IF(B{qf2}>=B{qs2},"Fast","Slow")')
    fm(ws, r, 4, f'=IF(B{r}=C{r},"✓ Converged","✗ Changed")')
    dat(ws, r, 1, 1); r += 1

# ── Section 5: π* from Q* ──
r += 1; sec(ws, r, "Section 5: Optimal Policy π*(s) = argmax_a Q*(s,a)")
r += 1
for j, h in enumerate(["State", "π*(s) ✱", "Q*(s, π*(s)) ✱"]): ws.cell(r, 1+j, value=h)
hdr(ws, r, 1, 3); r += 1
for st, qs, qf in [("Cool", q_star_rows["(Cool, Slow)"], q_star_rows["(Cool, Fast)"]),
                    ("Warm", q_star_rows["(Warm, Slow)"], q_star_rows["(Warm, Fast)"])]:
    ws.cell(r, 1, value=st); ws.cell(r, 1).font = Font(bold=True)
    fm(ws, r, 2, f'=IF(B{qf}>=B{qs},"Fast","Slow")')
    fm(ws, r, 3, f"=MAX(B{qs},B{qf})")
    dat(ws, r, 1, 1); r += 1
r_tmp = r
ws.cell(r, 1, value="Overheated"); ws.cell(r, 1).font = Font(bold=True)
ws.cell(r, 2, value="— (terminal)"); ws.cell(r, 3, value=0)
dat(ws, r, 1, 3)
for c in range(1, 4): ws.cell(r, c).fill = TF

aw(ws)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 5: MC — Episodes (trajectories + return-to-go)
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("MC-Episodes")
ttl(ws, 1, "Monte Carlo — Episodes with Return-to-Go", "G")
nt(ws, 2, "NOTE: t = time step within an episode (one single transition s→s').  Unlike DP's k (full sweep over all states), t advances one step at a time.", "G")
nt(ws, 3, "F=0.9^A (discount).  G_T=E_T (last step r).  G_t=E_t+0.9·G_{t+1} (backward recursion).", "G")

ep_hdr = ["Step (t)", "s", "a", "s'", "r", "γ^t ✱", "G_t ✱"]
cur = 5
fv_state = []
fv_sa = []

for ei, traj in enumerate(episodes):
    ws.merge_cells(f"A{cur}:G{cur}")
    ws.cell(cur, 1, value=f"Episode {ei+1}")
    ws.cell(cur, 1).font = Font(bold=True, size=12, color="FFFFFF"); ws.cell(cur, 1).alignment = CT
    for c in range(1, 8): ws.cell(cur, c).fill = EF
    cur += 1
    for j, h in enumerate(ep_hdr): ws.cell(cur, 1+j, value=h)
    hdr(ws, cur, 1, 7); cur += 1

    ds = cur
    vs, vsa = {}, {}
    for si, (t, s, a, sn, r) in enumerate(traj):
        ws.cell(cur,1,value=t); ws.cell(cur,2,value=s); ws.cell(cur,3,value=a)
        ws.cell(cur,4,value=sn); ws.cell(cur,5,value=r)
        fm(ws, cur, 6, f"=0.9^A{cur}")
        dat(ws, cur, 1, 5)
        if sn == "Overheated":
            for c in range(1, 6): ws.cell(cur, c).fill = TF
        if s not in vs: vs[s] = cur
        if (s, a) not in vsa: vsa[(s, a)] = cur
        cur += 1
    de = cur - 1

    fm(ws, de, 7, f"=E{de}")
    for row in range(de - 1, ds - 1, -1):
        fm(ws, row, 7, f"=E{row}+0.9*G{row+1}")

    fv_state.append(vs)
    fv_sa.append(vsa)

    ws.cell(cur, 1, value="G₀ ="); ws.cell(cur, 1).font = Font(bold=True)
    fm(ws, cur, 2, f"=G{ds}")
    cur += 2

aw(ws)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 6: MC — V and Q Estimation (First-Visit)
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("MC-VQ")
ttl(ws, 1, "Monte Carlo — First-Visit V(s) and Q(s,a) Estimation", "H")
nt(ws, 2, "Each cell = G_t of the first visit in that episode (cross-sheet ref).  V/Q = AVERAGE of those returns.", "H")

# ── V^MC ──
r = 4; sec(ws, r, "V^MC(s) = average of first-visit G_t per state")
r += 1
v_mc_hdrs = ["State", "Ep1 G_t ✱", "Ep2 G_t ✱", "Ep3 G_t ✱", "Ep4 G_t ✱", "Ep5 G_t ✱", "V^MC(s) ✱"]
for j, h in enumerate(v_mc_hdrs): ws.cell(r, 1+j, value=h)
hdr(ws, r, 1, 7); r += 1
v_mc_rows = {}
for st in ["Cool", "Warm"]:
    ws.cell(r, 1, value=st); ws.cell(r, 1).font = Font(bold=True)
    for ei in range(5):
        col = 2 + ei
        if st in fv_state[ei]:
            fm(ws, r, col, f"='MC-Episodes'!G{fv_state[ei][st]}")
    fm(ws, r, 7, f"=AVERAGE(B{r}:F{r})")
    dat(ws, r, 1, 1)
    v_mc_rows[st] = r; r += 1
ws.cell(r, 1, value="Overheated"); ws.cell(r, 1).font = Font(bold=True)
ws.cell(r, 2, value=0); ws.cell(r, 7, value=0)
dat(ws, r, 1, 7)
for c in range(1, 8): ws.cell(r, c).fill = TF
r += 1

# ── Q^MC ──
r += 1; sec(ws, r, "Q^MC(s,a) = average of first-visit G_t per (s,a)")
r += 1
q_mc_hdrs = ["(s, a)", "Ep1 G_t ✱", "Ep2 G_t ✱", "Ep3 G_t ✱", "Ep4 G_t ✱", "Ep5 G_t ✱", "Q^MC(s,a) ✱"]
for j, h in enumerate(q_mc_hdrs): ws.cell(r, 1+j, value=h)
hdr(ws, r, 1, 7); r += 1
q_mc_rows = {}
for s, a in [("Cool","Slow"),("Cool","Fast"),("Warm","Slow"),("Warm","Fast")]:
    label = f"({s}, {a})"
    ws.cell(r, 1, value=label); ws.cell(r, 1).font = Font(bold=True)
    for ei in range(5):
        col = 2 + ei
        if (s, a) in fv_sa[ei]:
            fm(ws, r, col, f"='MC-Episodes'!G{fv_sa[ei][(s, a)]}")
    fm(ws, r, 7, f"=AVERAGE(B{r}:F{r})")
    dat(ws, r, 1, 1)
    q_mc_rows[(s, a)] = r; r += 1

# ── Greedy from MC ──
r += 1; sec(ws, r, "Greedy Policy from Q^MC: π(s) = argmax_a Q^MC(s,a)")
r += 1
for j, h in enumerate(["State", "π(s) ✱", "Q^MC(s,π(s)) ✱"]): ws.cell(r, 1+j, value=h)
hdr(ws, r, 1, 3); r += 1
for st, qs_r, qf_r in [("Cool", q_mc_rows[("Cool","Slow")], q_mc_rows[("Cool","Fast")]),
                        ("Warm", q_mc_rows[("Warm","Slow")], q_mc_rows[("Warm","Fast")])]:
    ws.cell(r, 1, value=st); ws.cell(r, 1).font = Font(bold=True)
    fm(ws, r, 2, f'=IF(G{qf_r}>=G{qs_r},"Fast","Slow")')
    fm(ws, r, 3, f"=MAX(G{qs_r},G{qf_r})")
    dat(ws, r, 1, 1); r += 1

aw(ws)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 7: TD(0) — V(s) Estimation
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("TD0")
ttl(ws, 1, "TD(0) — Step-by-Step V(s) Updates", "M")
nt(ws, 2, "NOTE: Each row = one time step t (a single transition s→s'), NOT a full sweep.  TD updates V(s) after every step, unlike DP which sweeps all states per iteration k.", "M")
nt(ws, 3, "V(s) ← V(s) + α·[r + γ·V(s') − V(s)].  α=0.1, γ=0.9.  L,M track V(Cool),V(Warm) after each step.", "M")

r = 5
td_hdrs = ["Step","Ep","s","a","r","s'","V(s) bef ✱","V(s') ✱","Target ✱","Error ✱","V(s) aft ✱","V(Cool) ✱","V(Warm) ✱"]
for j, h in enumerate(td_hdrs): ws.cell(r, 1+j, value=h)
hdr(ws, r, 1, 13)

# Init row
r = 6; td_init = r
for c in range(1, 12): ws.cell(r, c, value="init" if c <= 6 else "")
ws.cell(r, 12, value=0); ws.cell(r, 13, value=0)
dat(ws, r, 1, 13); sub(ws, r, 12, 13)

for i, (ep, s, a, rew, sn, _) in enumerate(all_steps):
    r = td_init + 1 + i; p = r - 1
    ws.cell(r,1,value=i+1); ws.cell(r,2,value=ep); ws.cell(r,3,value=s)
    ws.cell(r,4,value=a); ws.cell(r,5,value=rew); ws.cell(r,6,value=sn)
    fm(ws, r, 7, f'=IF(C{r}="Cool",L{p},M{p})')
    fm(ws, r, 8, f'=IF(F{r}="Overheated",0,IF(F{r}="Cool",L{p},M{p}))')
    fm(ws, r, 9, f"=E{r}+0.9*H{r}")
    fm(ws, r, 10, f"=I{r}-G{r}")
    fm(ws, r, 11, f"=G{r}+0.1*J{r}")
    fm(ws, r, 12, f'=IF(C{r}="Cool",K{r},L{p})')
    fm(ws, r, 13, f'=IF(C{r}="Warm",K{r},M{p})')
    dat(ws, r, 1, 6)
    if sn == "Overheated":
        for c in range(1, 7): ws.cell(r, c).fill = TF

td_last = td_init + len(all_steps)

# Final V
r = td_last + 2
sec(ws, r, "Final V^TD(0)(s) after all episodes")
r += 1
for j, h in enumerate(["State", "V(s) ✱"]): ws.cell(r, 1+j, value=h)
hdr(ws, r, 1, 2); r += 1
ws.cell(r,1,value="Cool"); fm(ws, r, 2, f"=L{td_last}"); dat(ws, r, 1, 1); r += 1
ws.cell(r,1,value="Warm"); fm(ws, r, 2, f"=M{td_last}"); dat(ws, r, 1, 1); r += 1
ws.cell(r,1,value="Overheated"); ws.cell(r,2,value=0); dat(ws, r, 1, 2)
for c in [1,2]: ws.cell(r, c).fill = TF

aw(ws)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 8: SARSA — Q(s,a) Estimation (On-Policy)
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("SARSA")
ttl(ws, 1, "SARSA — On-Policy Q(s,a) Updates", "P")
nt(ws, 2, "NOTE: Each row = one time step t (a single transition s,a→r,s',a'), NOT a full sweep.  SARSA updates Q(s,a) after every step using the ACTUAL next action a'.", "P")
nt(ws, 3, "Q(s,a) ← Q(s,a) + α·[r + γ·Q(s',a') − Q(s,a)].  α=0.1, γ=0.9.  M-P track the full Q-table.", "P")

r = 5
sa_hdrs = ["Step","Ep","s","a","r","s'","a'","Q(s,a) bef ✱","Q(s',a') ✱","Target ✱","Error ✱","Q aft ✱",
           "Q(C,S) ✱","Q(C,F) ✱","Q(W,S) ✱","Q(W,F) ✱"]
for j, h in enumerate(sa_hdrs): ws.cell(r, 1+j, value=h)
hdr(ws, r, 1, 16)

r = 6; sa_init = r
for c in range(1, 13): ws.cell(r, c, value="init" if c <= 7 else "")
ws.cell(r,13,value=0); ws.cell(r,14,value=0); ws.cell(r,15,value=0); ws.cell(r,16,value=0)
dat(ws, r, 1, 16); sub(ws, r, 13, 16)

for i, (ep, s, a, rew, sn, an) in enumerate(all_steps):
    r = sa_init + 1 + i; p = r - 1
    ws.cell(r,1,value=i+1); ws.cell(r,2,value=ep); ws.cell(r,3,value=s)
    ws.cell(r,4,value=a); ws.cell(r,5,value=rew); ws.cell(r,6,value=sn); ws.cell(r,7,value=an)
    # H: Q(s,a) before
    fm(ws, r, 8,
       f'=IF(AND(C{r}="Cool",D{r}="Slow"),M{p},IF(AND(C{r}="Cool",D{r}="Fast"),N{p},IF(AND(C{r}="Warm",D{r}="Slow"),O{p},P{p})))')
    # I: Q(s',a')
    fm(ws, r, 9,
       f'=IF(F{r}="Overheated",0,IF(AND(F{r}="Cool",G{r}="Slow"),M{p},IF(AND(F{r}="Cool",G{r}="Fast"),N{p},IF(AND(F{r}="Warm",G{r}="Slow"),O{p},P{p}))))')
    fm(ws, r, 10, f"=E{r}+0.9*I{r}")      # Target
    fm(ws, r, 11, f"=J{r}-H{r}")            # Error
    fm(ws, r, 12, f"=H{r}+0.1*K{r}")        # Q after
    # M-P: Q-table state
    fm(ws, r, 13, f'=IF(AND(C{r}="Cool",D{r}="Slow"),L{r},M{p})')
    fm(ws, r, 14, f'=IF(AND(C{r}="Cool",D{r}="Fast"),L{r},N{p})')
    fm(ws, r, 15, f'=IF(AND(C{r}="Warm",D{r}="Slow"),L{r},O{p})')
    fm(ws, r, 16, f'=IF(AND(C{r}="Warm",D{r}="Fast"),L{r},P{p})')
    dat(ws, r, 1, 7)
    if sn == "Overheated":
        for c in range(1, 8): ws.cell(r, c).fill = TF

sa_last = sa_init + len(all_steps)

r = sa_last + 2; sec(ws, r, "Final Q^SARSA(s,a)")
r += 1
for j, h in enumerate(["State\\Action", "Slow ✱", "Fast ✱"]): ws.cell(r, 1+j, value=h)
hdr(ws, r, 1, 3); r += 1
ws.cell(r,1,value="Cool"); fm(ws, r, 2, f"=M{sa_last}"); fm(ws, r, 3, f"=N{sa_last}"); dat(ws, r, 1, 1); r += 1
ws.cell(r,1,value="Warm"); fm(ws, r, 2, f"=O{sa_last}"); fm(ws, r, 3, f"=P{sa_last}"); dat(ws, r, 1, 1); r += 1
ws.cell(r,1,value="Overheated"); ws.cell(r,2,value=0); ws.cell(r,3,value=0)
dat(ws, r, 1, 3)
for c in [1,2,3]: ws.cell(r, c).fill = TF
r += 2
sec(ws, r, "Greedy Policy from Q^SARSA")
r += 1
for j, h in enumerate(["State", "π(s) ✱"]): ws.cell(r, 1+j, value=h)
hdr(ws, r, 1, 2); r += 1
sr_qt_row = sa_last + 4
ws.cell(r,1,value="Cool"); fm(ws, r, 2, f'=IF(C{sr_qt_row}>=B{sr_qt_row},"Fast","Slow")'); dat(ws, r, 1, 1); r += 1
ws.cell(r,1,value="Warm"); fm(ws, r, 2, f'=IF(C{sr_qt_row+1}>=B{sr_qt_row+1},"Fast","Slow")'); dat(ws, r, 1, 1)

aw(ws)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 9: Q-Learning — Q*(s,a) Estimation (Off-Policy)
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("QLearning")
ttl(ws, 1, "Q-Learning — Off-Policy Q*(s,a) Updates", "O")
nt(ws, 2, "NOTE: Each row = one time step t (a single transition s,a→r,s').  Q-Learning uses max_{a'} Q(s',a') — the BEST possible action, not the one actually taken (off-policy).", "O")
nt(ws, 3, "Q(s,a) ← Q(s,a) + α·[r + γ·max_{a'} Q(s',a') − Q(s,a)].  α=0.1, γ=0.9.  L-O track Q-table.", "O")

r = 5
ql_hdrs = ["Step","Ep","s","a","r","s'","Q(s,a) bef ✱","max Q(s',·) ✱","Target ✱","Error ✱","Q aft ✱",
           "Q(C,S) ✱","Q(C,F) ✱","Q(W,S) ✱","Q(W,F) ✱"]
for j, h in enumerate(ql_hdrs): ws.cell(r, 1+j, value=h)
hdr(ws, r, 1, 15)

r = 6; ql_init = r
for c in range(1, 12): ws.cell(r, c, value="init" if c <= 6 else "")
ws.cell(r,12,value=0); ws.cell(r,13,value=0); ws.cell(r,14,value=0); ws.cell(r,15,value=0)
dat(ws, r, 1, 15); sub(ws, r, 12, 15)

for i, (ep, s, a, rew, sn, _) in enumerate(all_steps):
    r = ql_init + 1 + i; p = r - 1
    ws.cell(r,1,value=i+1); ws.cell(r,2,value=ep); ws.cell(r,3,value=s)
    ws.cell(r,4,value=a); ws.cell(r,5,value=rew); ws.cell(r,6,value=sn)
    # G: Q(s,a) before
    fm(ws, r, 7,
       f'=IF(AND(C{r}="Cool",D{r}="Slow"),L{p},IF(AND(C{r}="Cool",D{r}="Fast"),M{p},IF(AND(C{r}="Warm",D{r}="Slow"),N{p},O{p})))')
    # H: max_a' Q(s',a')
    fm(ws, r, 8,
       f'=IF(F{r}="Overheated",0,IF(F{r}="Cool",MAX(L{p},M{p}),MAX(N{p},O{p})))')
    fm(ws, r, 9, f"=E{r}+0.9*H{r}")       # Target
    fm(ws, r, 10, f"=I{r}-G{r}")            # Error
    fm(ws, r, 11, f"=G{r}+0.1*J{r}")        # Q after
    # L-O: Q-table state
    fm(ws, r, 12, f'=IF(AND(C{r}="Cool",D{r}="Slow"),K{r},L{p})')
    fm(ws, r, 13, f'=IF(AND(C{r}="Cool",D{r}="Fast"),K{r},M{p})')
    fm(ws, r, 14, f'=IF(AND(C{r}="Warm",D{r}="Slow"),K{r},N{p})')
    fm(ws, r, 15, f'=IF(AND(C{r}="Warm",D{r}="Fast"),K{r},O{p})')
    dat(ws, r, 1, 6)
    if sn == "Overheated":
        for c in range(1, 7): ws.cell(r, c).fill = TF

ql_last = ql_init + len(all_steps)

r = ql_last + 2; sec(ws, r, "Final Q^QLearning(s,a)")
r += 1
for j, h in enumerate(["State\\Action", "Slow ✱", "Fast ✱"]): ws.cell(r, 1+j, value=h)
hdr(ws, r, 1, 3); r += 1
ql_qt_row = r
ws.cell(r,1,value="Cool"); fm(ws, r, 2, f"=L{ql_last}"); fm(ws, r, 3, f"=M{ql_last}"); dat(ws, r, 1, 1); r += 1
ws.cell(r,1,value="Warm"); fm(ws, r, 2, f"=N{ql_last}"); fm(ws, r, 3, f"=O{ql_last}"); dat(ws, r, 1, 1); r += 1
ws.cell(r,1,value="Overheated"); ws.cell(r,2,value=0); ws.cell(r,3,value=0)
dat(ws, r, 1, 3)
for c in [1,2,3]: ws.cell(r, c).fill = TF
r += 2
sec(ws, r, "Greedy Policy from Q-Learning")
r += 1
for j, h in enumerate(["State", "π*(s) ✱", "V*(s) = max Q ✱"]): ws.cell(r, 1+j, value=h)
hdr(ws, r, 1, 3); r += 1
ws.cell(r,1,value="Cool"); ws.cell(r,1).font = Font(bold=True)
fm(ws, r, 2, f'=IF(C{ql_qt_row}>=B{ql_qt_row},"Fast","Slow")')
fm(ws, r, 3, f"=MAX(B{ql_qt_row},C{ql_qt_row})")
dat(ws, r, 1, 1); r += 1
ws.cell(r,1,value="Warm"); ws.cell(r,1).font = Font(bold=True)
fm(ws, r, 2, f'=IF(C{ql_qt_row+1}>=B{ql_qt_row+1},"Fast","Slow")')
fm(ws, r, 3, f"=MAX(B{ql_qt_row+1},C{ql_qt_row+1})")
dat(ws, r, 1, 1)

aw(ws)


# ══════════════════════════════════════════════════════════════════════════════
# Save
# ══════════════════════════════════════════════════════════════════════════════
out = "/Users/samrat.kar/git/drl/1-introduction/assets/mdp_race_car.xlsx"
wb.save(out)
print(f"Saved: {out}")
print(f"Sheets: {wb.sheetnames}")
print(f"Episodes: {[len(e) for e in episodes]} steps each, {len(all_steps)} total steps")
